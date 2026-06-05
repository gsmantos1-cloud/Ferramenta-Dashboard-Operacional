import turso_db as sqlite3
import threading
import os
import csv
import json
import re
import urllib.parse
import urllib.request
from datetime import date, timedelta, datetime
from functools import wraps
from io import BytesIO, StringIO

import holidays
import openpyxl
from flask import Flask, jsonify, render_template, request, send_file, session, redirect, url_for
from werkzeug.security import check_password_hash

try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "dev-secret-key-change-me")

LOGIN_USUARIO    = os.getenv("LOGIN_USUARIO", "gs.operacional")
LOGIN_SENHA_HASH = os.getenv("LOGIN_SENHA_HASH", "")
LOGIN_SENHA      = os.getenv("LOGIN_SENHA", "")   # senha em texto plano (alternativa ao hash)


@app.after_request
def add_ngrok_header(response):
    response.headers["ngrok-skip-browser-warning"] = "true"
    return response


@app.context_processor
def inject_globals():
    logo_path = os.path.join(BASE_DIR, "static", "logo.png")
    return {"logo_exists": os.path.exists(logo_path)}


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logado"):
            if request.path.startswith("/api/"):
                return jsonify({"erro": "Não autorizado"}), 401
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return decorated

# ── Alertas de estoque (polling — SSE removido para compatibilidade Vercel) ──
def push_alert(data: dict):
    """No Vercel, alertas são consultados via polling. Esta função é no-op."""
    pass

def _verificar_e_alertar_estoque(conn, numero_pedido: str, cliente: str, produtos: list):
    """Checa estoque dos itens de um pedido recém-salvo e dispara alertas SSE."""
    for p in produtos:
        vid       = p.get("variant_id")
        nome      = p.get("name", "Produto")
        var_vals  = p.get("variant_values") or []
        variante  = ", ".join(str(v) for v in var_vals) if var_vals else ""
        qtd_ped   = int(p.get("quantity") or 1)
        if not vid:
            continue
        stock = conn.execute(
            "SELECT quantity, min_quantity, sku FROM sku_stock WHERE nv_variant_id = ?", (vid,)
        ).fetchone()
        if not stock:
            continue
        qty     = stock["quantity"]
        min_qty = stock["min_quantity"]
        if qty <= 0:
            status_est = "sem_estoque"
        elif qty <= min_qty:
            status_est = "estoque_baixo"
        else:
            continue  # estoque OK — sem alerta
        push_alert({
            "type":           "alerta_venda",
            "status":         status_est,
            "numero_pedido":  numero_pedido,
            "cliente":        cliente,
            "produto":        nome,
            "variante":       variante,
            "sku":            stock["sku"] or "",
            "qty_atual":      qty,
            "qty_pedida":     qtd_ped,
            "min_qty":        min_qty,
            "timestamp":      datetime.now().strftime("%H:%M:%S"),
        })

STATUS_ORDER = ["NO PRAZO", "Atraso leve", "Atraso moderado", "Atraso crítico"]
PERS_STATUS  = ["A SEPARAR", "SEPARAÇÃO", "NA PERSONALIZAÇÃO", "PRONTO", "ENVIADO"]

LIMITES = {
    "Normal": [
        (0, 3,    "NO PRAZO"),
        (4, 5,    "Atraso leve"),
        (6, 6,    "Atraso moderado"),
        (7, 9999, "Atraso crítico"),
    ],
    "Personalizado": [
        (0, 4,    "NO PRAZO"),
        (5, 6,    "Atraso leve"),
        (7, 7,    "Atraso moderado"),
        (8, 9999, "Atraso crítico"),
    ],
    "Internacional": [
        (0, 15,   "NO PRAZO"),
        (16, 20,  "Atraso leve"),
        (21, 25,  "Atraso moderado"),
        (26, 9999,"Atraso crítico"),
    ],
}


# ── Database ─────────────────────────────────────────────────────────────────

def init_db():
    with get_conn() as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS romaneios (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                criado_em   TEXT    NOT NULL DEFAULT (datetime('now','localtime')),
                total       INTEGER NOT NULL DEFAULT 0,
                observacao  TEXT
            );

            CREATE TABLE IF NOT EXISTS config (
                chave TEXT PRIMARY KEY,
                valor TEXT
            );

            CREATE TABLE IF NOT EXISTS personalizacoes (
                id                    INTEGER PRIMARY KEY AUTOINCREMENT,
                numero_pedido         TEXT    NOT NULL,
                nome_personalizacao   TEXT,
                numero_personalizacao TEXT,
                status                TEXT    NOT NULL DEFAULT 'A SEPARAR',
                observacao            TEXT,
                criado_em             TEXT    DEFAULT (datetime('now','localtime')),
                atualizado_em         TEXT    DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS pedidos (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                numero           TEXT    NOT NULL,
                data_pedido      TEXT    NOT NULL,
                categoria        TEXT    NOT NULL DEFAULT 'Normal',
                status           TEXT    NOT NULL DEFAULT 'NO PRAZO',
                suspeito         INTEGER DEFAULT 0,
                cliente          TEXT,
                total            TEXT,
                pagamento        TEXT,
                forma_pagamento  TEXT,
                transportadora   TEXT,
                criado_em        TEXT    DEFAULT (datetime('now','localtime')),
                enviado_em       TEXT    NULL,
                status_ao_enviar TEXT    NULL,
                romaneio_id      INTEGER REFERENCES romaneios(id),
                ativo            INTEGER DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS pedido_itens (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                pedido_numero TEXT NOT NULL,
                produto_nome  TEXT,
                variante      TEXT,
                quantidade    INTEGER DEFAULT 1,
                preco_unit    TEXT
            );

            CREATE TABLE IF NOT EXISTS sku_costs (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                sku            TEXT    NOT NULL,
                name           TEXT,
                type           TEXT    NOT NULL DEFAULT 'product'
                                       CHECK(type IN ('product','brinde')),
                cost           REAL    NOT NULL CHECK(cost >= 0),
                effective_from TEXT    NOT NULL,
                effective_to   TEXT,
                notes          TEXT,
                created_at     DATETIME DEFAULT CURRENT_TIMESTAMP,
                nv_variant_id  INTEGER,
                nv_product_id  INTEGER
            );

            CREATE TABLE IF NOT EXISTS sku_stock (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                nv_variant_id   INTEGER UNIQUE,
                nv_product_id   INTEGER,
                sku             TEXT,
                quantity        INTEGER NOT NULL DEFAULT 0,
                min_quantity    INTEGER NOT NULL DEFAULT 3,
                updated_at      DATETIME DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS sku_stock_movements (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                nv_variant_id   INTEGER,
                sku             TEXT,
                tipo            TEXT    NOT NULL
                                        CHECK(tipo IN ('entrada','saida_venda','saida_manual','ajuste')),
                quantidade      INTEGER NOT NULL,
                pedido_numero   TEXT,
                observacao      TEXT,
                created_at      DATETIME DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS sku_pers_pricing (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                nv_variant_id   INTEGER UNIQUE,
                sku             TEXT,
                custo_nome      REAL    NOT NULL DEFAULT 0,
                custo_numero    REAL    NOT NULL DEFAULT 0,
                custo_escudo    REAL    NOT NULL DEFAULT 0,
                created_at      DATETIME DEFAULT (datetime('now','localtime')),
                updated_at      DATETIME DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS atacado_pedidos (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                cliente    TEXT NOT NULL,
                contato    TEXT,
                endereco   TEXT,
                observacao TEXT,
                prazo      TEXT,
                status     TEXT NOT NULL DEFAULT 'pendente',
                criado_por TEXT,
                created_at DATETIME DEFAULT (datetime('now','localtime')),
                updated_at DATETIME DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS atacado_itens (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                pedido_id    INTEGER NOT NULL REFERENCES atacado_pedidos(id) ON DELETE CASCADE,
                produto      TEXT NOT NULL,
                variante     TEXT,
                quantidade   INTEGER NOT NULL DEFAULT 1,
                separado     INTEGER NOT NULL DEFAULT 0,
                separado_em  DATETIME,
                separado_por TEXT
            );

            CREATE TABLE IF NOT EXISTS compras_manual (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                nome        TEXT    NOT NULL,
                sku         TEXT,
                qty_comprar INTEGER NOT NULL DEFAULT 1,
                custo_unit  REAL    NOT NULL DEFAULT 0,
                observacao  TEXT,
                data        TEXT    NOT NULL DEFAULT (date('now','localtime')),
                created_at  DATETIME DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS compras_registros (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                data         TEXT    NOT NULL DEFAULT (date('now','localtime')),
                produto_nome TEXT    NOT NULL,
                nv_product_id INTEGER,
                fornecedor   TEXT,
                preco_unit   REAL    NOT NULL DEFAULT 0,
                observacao   TEXT,
                criado_por   TEXT,
                created_at   DATETIME DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS compras_tamanhos (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                compra_id  INTEGER NOT NULL REFERENCES compras_registros(id) ON DELETE CASCADE,
                tamanho    TEXT    NOT NULL,
                quantidade INTEGER NOT NULL DEFAULT 0
            );
        """)
        for col, coltype in [
            ("cliente", "TEXT"), ("total", "TEXT"), ("pagamento", "TEXT"),
            ("forma_pagamento", "TEXT"), ("transportadora", "TEXT"),
            ("suspeito", "INTEGER"), ("romaneio_id", "INTEGER"),
        ]:
            try:
                conn.execute(f'ALTER TABLE pedidos ADD COLUMN {col} {coltype}')
            except Exception:
                pass

        # atacado migrations (idempotent)
        for col in [
            "cep TEXT", "cpf TEXT", "cidade TEXT", "nome TEXT",
            "pago INTEGER DEFAULT 0",
        ]:
            try: conn.execute(f"ALTER TABLE atacado_pedidos ADD COLUMN {col}")
            except Exception: pass

        for col in [
            "nv_variant_id INTEGER",
            "qty_estoque INTEGER DEFAULT 0",
            "qty_fornecedor INTEGER DEFAULT 0",
            "estoque_descontado INTEGER DEFAULT 0",
            "valor_unit REAL DEFAULT 0",
        ]:
            try: conn.execute(f"ALTER TABLE atacado_itens ADD COLUMN {col}")
            except Exception: pass

        for col in ["frete_tipo TEXT DEFAULT 'a_combinar'", "numero INTEGER"]:
            try: conn.execute(f"ALTER TABLE atacado_pedidos ADD COLUMN {col}")
            except Exception: pass

        # Backfill: numera pedidos antigos sem número (ordem por id)
        try:
            sem_numero = conn.execute(
                "SELECT id FROM atacado_pedidos WHERE numero IS NULL ORDER BY id"
            ).fetchall()
            if sem_numero:
                maxn = conn.execute(
                    "SELECT COALESCE(MAX(numero), 0) FROM atacado_pedidos"
                ).fetchone()[0] or 0
                n = maxn
                for row in sem_numero:
                    n += 1
                    conn.execute("UPDATE atacado_pedidos SET numero=? WHERE id=?", (n, row[0]))
        except Exception:
            pass

        # sku_stock — nomes para sync
        for col in ["produto_nome TEXT", "variante_label TEXT"]:
            try: conn.execute(f"ALTER TABLE sku_stock ADD COLUMN {col}")
            except Exception: pass

        conn.commit()

        # sku_costs migrations (idempotent)
        for col in ["nv_variant_id INTEGER", "nv_product_id INTEGER"]:
            try:
                conn.execute(f'ALTER TABLE sku_costs ADD COLUMN {col}')
            except Exception:
                pass

        # sku_stock_movements — preço de compra por lote (para CMP)
        try: conn.execute("ALTER TABLE sku_stock_movements ADD COLUMN preco_compra REAL")
        except Exception: pass

        # pedido_itens migrations (idempotent)
        for col in ["imagem_url TEXT", "nv_variant_id INTEGER"]:
            try:
                conn.execute(f'ALTER TABLE pedido_itens ADD COLUMN {col}')
            except Exception:
                pass

        # pedidos: controle de processamento de estoque
        # Default 0 = não processado; pedidos existentes recebem 1 para não descontar retroativamente
        try:
            conn.execute('ALTER TABLE pedidos ADD COLUMN estoque_processado INTEGER DEFAULT 0')
            conn.execute('UPDATE pedidos SET estoque_processado = 1')  # histórico já existente = não processar
        except Exception:
            pass


def get_conn():
    return sqlite3.connect()


# ── Business logic ────────────────────────────────────────────────────────────

def calcular_dias_uteis(data_inicio: date, data_fim: date) -> int:
    feriados = holidays.Brazil(years=range(data_inicio.year, data_fim.year + 2))
    count, atual = 0, data_inicio
    while atual <= data_fim:
        if atual.weekday() < 5 and atual not in feriados:
            count += 1
        atual += timedelta(days=1)
    return count


def determinar_status(dias: int, categoria: str) -> str:
    for minimo, maximo, status in LIMITES.get(categoria, LIMITES["Normal"]):
        if minimo <= dias <= maximo:
            return status
    return "Atraso crítico"


def atualizar_status_db():
    hoje = date.today()
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT id, data_pedido, categoria, status FROM pedidos WHERE ativo = 1"
        ).fetchall()
        for r in rows:
            dp   = date.fromisoformat(r["data_pedido"])
            dias = calcular_dias_uteis(dp, hoje)
            novo = determinar_status(dias, r["categoria"])
            if novo != r["status"]:
                conn.execute("UPDATE pedidos SET status = ? WHERE id = ?", (novo, r["id"]))


def periodo_para_data_inicio(periodo: str) -> date:
    hoje = date.today()
    if periodo == "diario":
        return hoje
    if periodo == "semanal":
        return hoje - timedelta(days=7)
    if periodo == "quinzenal":
        return hoje - timedelta(days=15)
    if periodo == "30dias":
        return hoje - timedelta(days=30)
    if periodo == "mes_atual":
        return hoje.replace(day=1)
    if periodo == "3meses":
        mes = hoje.month - 3
        ano = hoje.year
        if mes <= 0:
            mes += 12
            ano -= 1
        return date(ano, mes, 1)
    return hoje - timedelta(days=30)


# ── Routes — auth ─────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login_page():
    if session.get("logado"):
        return redirect(url_for("dashboard"))
    erro = None
    if request.method == "POST":
        usuario = request.form.get("usuario", "").strip()
        senha   = request.form.get("senha", "")
        senha_ok = (
            (LOGIN_SENHA and senha == LOGIN_SENHA) or
            (LOGIN_SENHA_HASH and check_password_hash(LOGIN_SENHA_HASH, senha))
        )
        if usuario == LOGIN_USUARIO and senha_ok:
            session["logado"]  = True
            session["usuario"] = usuario
            return redirect(url_for("dashboard"))
        erro = "Usuário ou senha incorretos."
    return render_template("login.html", erro=erro)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login_page"))


# ── Routes — pages ────────────────────────────────────────────────────────────

@app.route("/")
@login_required
def dashboard():
    return render_template("dashboard.html")


@app.route("/relatorios")
@login_required
def relatorios():
    return render_template("relatorios.html")


@app.route("/relatorio/vendas-pdf")
@login_required
def relatorio_vendas_pdf():
    return render_template("relatorio_vendas.html")


@app.route("/api/relatorio/vendas-excel")
@login_required
def api_relatorio_vendas_excel():
    import re as _re, io
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter

    inicio = request.args.get("inicio", "")
    fim    = request.args.get("fim", "")

    with get_conn() as conn:
        filtro_data = ""
        params: list = []
        if inicio and fim:
            filtro_data = "AND date(p.data_pedido) >= ? AND date(p.data_pedido) <= ?"
            params = [inicio, fim]

        # ── Dados detalhados por variante + data ──────────────────────────
        rows = conn.execute(f"""
            SELECT pi.produto_nome,
                   pi.variante,
                   pi.quantidade,
                   pi.preco_unit,
                   p.data_pedido,
                   p.numero AS pedido_numero
            FROM pedido_itens pi
            JOIN pedidos p ON p.numero = pi.pedido_numero
            WHERE p.ativo IN (0, 1)
            {filtro_data}
            ORDER BY p.data_pedido ASC, pi.produto_nome ASC
        """, params).fetchall()

    dados = [dict(r) for r in rows]

    is_brinde = lambda r: not r["preco_unit"] or str(r["preco_unit"]) in ("None", "0", "0.0", "0.00")

    def extrair_tamanho(v):
        if not v: return ""
        for parte in [x.strip() for x in v.split(",")]:
            if _re.match(r'^(PP?|M|G{1,3}|XG{1,2}|\d{1,3})$', parte, _re.I):
                return parte.upper()
        return v.split(",")[0].strip()

    # ── Agrupa por produto para aba de resumo ─────────────────────────────
    resumo: dict = {}
    for r in dados:
        nome_raw  = r["produto_nome"] or ""
        nome_base = _re.sub(r"\s*\([^)]+\)\s*$", "", nome_raw).strip() or nome_raw
        tipo      = "Brinde" if is_brinde(r) else "Produto"
        key       = (nome_base, tipo)
        if key not in resumo:
            resumo[key] = {"produto": nome_base, "tipo": tipo, "total": 0, "tamanhos": {}}
        qty = r["quantidade"] or 0
        resumo[key]["total"] += qty
        tam = extrair_tamanho(r["variante"])
        if tam:
            resumo[key]["tamanhos"][tam] = resumo[key]["tamanhos"].get(tam, 0) + qty

    ORDEM_TAM = {"PP":0,"P":1,"M":2,"G":3,"GG":4,"GGG":5,"XG":6,"XGG":7}
    resumo_list = sorted(resumo.values(), key=lambda x: (-x["total"], x["tipo"]))

    # ── Estilos ────────────────────────────────────────────────────────────
    HDR_FILL    = PatternFill("solid", fgColor="1A3A5C")   # azul escuro
    HDR_FONT    = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    TITLE_FONT  = Font(name="Calibri", bold=True, size=14, color="1A3A5C")
    SUB_FONT    = Font(name="Calibri", size=9, color="666666")
    PROD_FILL   = PatternFill("solid", fgColor="EBF3FB")   # azul clarinho
    BRIN_FILL   = PatternFill("solid", fgColor="FEF3E2")   # laranja clarinho
    ALT_FILL    = PatternFill("solid", fgColor="F7F7F7")
    CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    RIGHT       = Alignment(horizontal="right",  vertical="center")
    thin        = Side(style="thin", color="D0D0D0")
    BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)
    BORD_TOP    = Border(top=Side(style="medium", color="1A3A5C"))
    GREEN_FONT  = Font(name="Calibri", bold=True, color="1A7A3C", size=11)
    ORANGE_FONT = Font(name="Calibri", bold=True, color="C05A00", size=11)

    wb = Workbook()

    # ══════════════════════════════════════════════════════════
    # ABA 1 — RESUMO POR PRODUTO
    # ══════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Resumo por Produto"
    ws1.sheet_view.showGridLines = False

    periodo_txt = f"{inicio} a {fim}" if inicio and fim else "Todos os registros"

    # Título
    ws1.merge_cells("A1:F1")
    ws1["A1"] = "GS Mantos — Relatório de Vendas por Produto"
    ws1["A1"].font = TITLE_FONT
    ws1["A1"].alignment = LEFT
    ws1.merge_cells("A2:F2")
    ws1["A2"] = f"Período: {periodo_txt}   |   Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws1["A2"].font = SUB_FONT
    ws1["A2"].alignment = LEFT
    ws1.row_dimensions[1].height = 22
    ws1.row_dimensions[2].height = 16

    headers1 = ["#", "Produto", "Tamanhos Vendidos", "Total", "% do Grupo"]
    COLS1 = 5

    total_geral   = sum(r["total"] for r in resumo_list)
    produtos_list = [r for r in resumo_list if r["tipo"] == "Produto"]
    brindes_list  = [r for r in resumo_list if r["tipo"] == "Brinde"]
    total_prod    = sum(r["total"] for r in produtos_list)
    total_brin    = sum(r["total"] for r in brindes_list)

    SEC_PROD_FILL  = PatternFill("solid", fgColor="1A3A5C")   # azul escuro — seção produto
    SEC_BRIN_FILL  = PatternFill("solid", fgColor="7A3A00")   # laranja escuro — seção brinde
    SEC_FONT       = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    TOTAL_FILL     = PatternFill("solid", fgColor="D9E8F5")
    TOTAL_FILL_B   = PatternFill("solid", fgColor="FAE5C8")
    TOTAL_FONT     = Font(name="Calibri", bold=True, size=10, color="1A3A5C")
    TOTAL_FONT_B   = Font(name="Calibri", bold=True, size=10, color="7A3A00")

    def write_section(ws, start_row, titulo, lista, total_sec, row_fill, alt_fill,
                      sec_fill, sec_font, total_fill, total_font_style, num_color):
        """Escreve uma seção (Produtos ou Brindes) na planilha."""
        # ── Cabeçalho da seção ──
        ws.merge_cells(f"A{start_row}:E{start_row}")
        cell = ws.cell(row=start_row, column=1, value=titulo)
        cell.font   = sec_font
        cell.fill   = sec_fill
        cell.alignment = LEFT
        ws.row_dimensions[start_row].height = 22
        start_row += 1

        # ── Cabeçalho de colunas ──
        for ci, h in enumerate(headers1, 1):
            c = ws.cell(row=start_row, column=ci, value=h)
            c.font = HDR_FONT; c.fill = HDR_FILL
            c.alignment = CENTER; c.border = BORDER
        ws.row_dimensions[start_row].height = 18
        start_row += 1

        # ── Linhas de dados ──
        for ri, item in enumerate(lista):
            row_num = start_row + ri
            fill    = row_fill if ri % 2 == 0 else alt_fill
            tam_str = "  ".join(
                f"{t} ×{q}" for t, q in
                sorted(item["tamanhos"].items(), key=lambda x: ORDEM_TAM.get(x[0], 99))
            ) or "—"
            pct = round(item["total"] / total_sec * 100, 1) if total_sec else 0
            vals = [ri+1, item["produto"], tam_str, item["total"], f"{pct}%"]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row_num, column=ci, value=v)
                c.fill = fill; c.border = BORDER
                c.alignment = CENTER if ci in (1,4,5) else LEFT
                if ci == 4:
                    c.font = Font(name="Calibri", bold=True, size=11, color=num_color)
            ws.row_dimensions[row_num].height = 18

        start_row += len(lista)

        # ── Linha de total da seção ──
        ws.merge_cells(f"A{start_row}:C{start_row}")
        # Apenas coluna 1 pode ter valor em células mescladas
        c = ws.cell(row=start_row, column=1, value=f"TOTAL {titulo.upper()}")
        c.font = total_font_style; c.fill = total_fill; c.alignment = LEFT; c.border = BORDER
        # Colunas B e C são MergedCell — aplicar apenas fill/border via referência direta é inválido
        # openpyxl gera automaticamente o merge, não tocar nas células filhas
        c = ws.cell(row=start_row, column=4, value=total_sec)
        c.font = total_font_style; c.fill = total_fill; c.alignment = CENTER; c.border = BORDER
        pct_tot = round(total_sec / total_geral * 100, 1) if total_geral else 0
        c = ws.cell(row=start_row, column=5, value=f"{pct_tot}% do total geral")
        c.font = total_font_style; c.fill = total_fill; c.alignment = CENTER; c.border = BORDER
        ws.row_dimensions[start_row].height = 20

        return start_row + 2   # pula uma linha em branco

    cur_row = 4
    cur_row = write_section(
        ws1, cur_row,
        titulo="🛍  Produtos Vendidos",
        lista=produtos_list, total_sec=total_prod,
        row_fill=PROD_FILL, alt_fill=ALT_FILL,
        sec_fill=SEC_PROD_FILL, sec_font=SEC_FONT,
        total_fill=TOTAL_FILL, total_font_style=TOTAL_FONT,
        num_color="1A3A5C",
    )
    cur_row = write_section(
        ws1, cur_row,
        titulo="🎁  Brindes Enviados",
        lista=brindes_list, total_sec=total_brin,
        row_fill=BRIN_FILL, alt_fill=PatternFill("solid", fgColor="FDF0DC"),
        sec_fill=SEC_BRIN_FILL, sec_font=SEC_FONT,
        total_fill=TOTAL_FILL_B, total_font_style=TOTAL_FONT_B,
        num_color="7A3A00",
    )

    # Larguras
    for col, w in zip("ABCDE", [6, 45, 32, 12, 13]):
        ws1.column_dimensions[get_column_letter(ord(col)-64)].width = w

    # ══════════════════════════════════════════════════════════
    # ABA 2 — VENDAS POR DIA
    # ══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Vendas por Dia")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:G1")
    ws2["A1"] = "GS Mantos — Vendas Detalhadas por Dia"
    ws2["A1"].font = TITLE_FONT
    ws2["A1"].alignment = LEFT
    ws2.merge_cells("A2:G2")
    ws2["A2"] = f"Período: {periodo_txt}   |   Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws2["A2"].font = SUB_FONT
    ws2["A2"].alignment = LEFT
    ws2.row_dimensions[1].height = 22
    ws2.row_dimensions[2].height = 16

    headers2 = ["Data", "Nº Pedido", "Tipo", "Produto", "Tamanho/Variante", "Qtd.", "Preço Unit."]
    for ci, h in enumerate(headers2, 1):
        cell = ws2.cell(row=4, column=ci, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    ws2.row_dimensions[4].height = 20

    # Ordena por data DESC + produto
    dados_ordenados = sorted(dados, key=lambda x: (x["data_pedido"] or "", x["produto_nome"] or ""), reverse=True)
    ultimo_dia = None
    for ri, r in enumerate(dados_ordenados):
        row_num  = ri + 5
        dia      = r["data_pedido"] or ""
        is_b     = is_brinde(r)
        fill     = BRIN_FILL if is_b else (PROD_FILL if ri % 2 == 0 else ALT_FILL)
        nome_raw = r["produto_nome"] or ""
        nome_base = _re.sub(r"\s*\([^)]+\)\s*$", "", nome_raw).strip() or nome_raw
        tam      = extrair_tamanho(r["variante"]) or (r["variante"] or "—")

        try:
            preco = float(str(r["preco_unit"]).replace(",","."))
        except Exception:
            preco = 0.0

        # Formata data pt-BR
        try:
            from datetime import date as _date
            d_obj = _date.fromisoformat(dia)
            dia_fmt = d_obj.strftime("%d/%m/%Y")
        except Exception:
            dia_fmt = dia

        vals = [dia_fmt, r["pedido_numero"], "Brinde" if is_b else "Produto",
                nome_base, tam, r["quantidade"] or 0, preco if preco else "—"]
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(row=row_num, column=ci, value=v)
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = CENTER if ci in (1,2,3,6) else LEFT
            if ci == 1 and dia != ultimo_dia:
                cell.font = Font(name="Calibri", bold=True, size=10, color="1A3A5C")
        ultimo_dia = dia
        ws2.row_dimensions[row_num].height = 17

    for col, w in zip("ABCDEFG", [12, 11, 9, 42, 22, 7, 12]):
        ws2.column_dimensions[get_column_letter(ord(col)-64)].width = w

    # ── Envia arquivo ──────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"vendas_gsmantos_{inicio or 'all'}_{fim or 'all'}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/relatorio/vendas-por-produto")
@login_required
def api_relatorio_vendas_produto():
    import re as _re
    inicio = request.args.get("inicio", "")
    fim    = request.args.get("fim", "")

    with get_conn() as conn:
        filtro_data = ""
        params: list = []
        if inicio and fim:
            filtro_data = "AND date(p.data_pedido) >= ? AND date(p.data_pedido) <= ?"
            params = [inicio, fim]

        rows = conn.execute(f"""
            SELECT pi.produto_nome,
                   pi.variante,
                   pi.imagem_url,
                   pi.preco_unit,
                   SUM(pi.quantidade) AS total_qty
            FROM pedido_itens pi
            JOIN pedidos p ON p.numero = pi.pedido_numero
            WHERE p.ativo IN (0, 1)
            {filtro_data}
            GROUP BY pi.produto_nome
            ORDER BY total_qty DESC
        """, params).fetchall()

    # Query detalhada por variante para calcular tamanhos
    rows_var = conn.execute(f"""
        SELECT pi.produto_nome,
               pi.variante,
               pi.imagem_url,
               pi.preco_unit,
               SUM(pi.quantidade) AS total_qty
        FROM pedido_itens pi
        JOIN pedidos p ON p.numero = pi.pedido_numero
        WHERE p.ativo IN (0, 1)
        {filtro_data}
        GROUP BY pi.produto_nome, pi.variante
        ORDER BY total_qty DESC
    """, params).fetchall()

    def _extrair_tamanho(variante_str):
        """Extrai o tamanho (P/M/G/GG/XG...) do campo variante."""
        if not variante_str:
            return None
        # Pega o primeiro trecho antes da vírgula e verifica se parece tamanho
        partes = [x.strip() for x in variante_str.split(",")]
        for p in partes:
            if _re.match(r'^(P|M|G{1,3}|XG|XGG|PP|GG|GGG|\d{1,3}(?:cm|ml)?)$', p, _re.IGNORECASE):
                return p.upper()
        return partes[0] if partes else None

    def _agrupar(lista):
        agrupado: dict = {}
        for r in lista:
            nome_raw  = r["produto_nome"] or ""
            nome_base = _re.sub(r"\s*\([^)]+\)\s*$", "", nome_raw).strip() or nome_raw

            if nome_base not in agrupado:
                agrupado[nome_base] = {
                    "produto_nome": nome_base,
                    "imagem_url":   r["imagem_url"] or "",
                    "total_qty":    0,
                    "tamanhos":     {},   # {tamanho: qty}
                }
            entry = agrupado[nome_base]
            qty = r["total_qty"] or 0
            entry["total_qty"] += qty
            if not entry["imagem_url"] and r["imagem_url"]:
                entry["imagem_url"] = r["imagem_url"]

            # Acumula por tamanho
            tam = _extrair_tamanho(r["variante"])
            if tam:
                entry["tamanhos"][tam] = entry["tamanhos"].get(tam, 0) + qty

        # Converte tamanhos dict → lista ordenada por qty desc
        ORDEM = {"P":0,"M":1,"G":2,"GG":3,"GGG":4,"XG":5,"XGG":6,"PP":7}
        for entry in agrupado.values():
            entry["tamanhos"] = sorted(
                [{"tamanho": k, "qty": v} for k, v in entry["tamanhos"].items()],
                key=lambda x: ORDEM.get(x["tamanho"].upper(), 99)
            )
        return sorted(agrupado.values(), key=lambda x: -x["total_qty"])

    # Separa brindes (preco_unit NULL ou '0' ou 'None') de produtos
    is_brinde = lambda r: not r["preco_unit"] or str(r["preco_unit"]) in ("None", "0", "0.0", "0.00")
    produtos_rows = [r for r in rows_var if not is_brinde(r)]
    brindes_rows  = [r for r in rows_var if     is_brinde(r)]

    return jsonify({
        "produtos": _agrupar(produtos_rows),
        "brindes":  _agrupar(brindes_rows),
    })


@app.route("/romaneio")
@login_required
def romaneio_page():
    return render_template("romaneio.html")


@app.route("/personalizacoes")
@login_required
def personalizacoes_page():
    return render_template("personalizacoes.html")


# ── Personalizações sync helper ───────────────────────────────────────────────

def _build_item_stmts(numero, produtos):
    """Monta os statements de INSERT dos itens de um pedido (para execução em batch)."""
    stmts = []
    for p in produtos:
        nome          = p.get("name", "")
        qtd           = p.get("quantity", 1)
        preco         = str(p.get("price", ""))
        nv_variant_id = p.get("variant_id")
        variante_vals = p.get("variant_values") or []
        variante      = ", ".join(str(v) for v in variante_vals) if variante_vals else None
        img_obj       = p.get("image") or {}
        imagem_url    = img_obj.get("src") or img_obj.get("url") or None
        if not imagem_url:
            imgs = p.get("images") or []
            if imgs:
                imagem_url = (imgs[0] or {}).get("src") or (imgs[0] or {}).get("url") or None
        stmts.append((
            """INSERT INTO pedido_itens
               (pedido_numero, produto_nome, variante, quantidade, preco_unit, imagem_url, nv_variant_id)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (numero, nome, variante, qtd, preco, imagem_url, nv_variant_id),
        ))
    return stmts


def _salvar_itens_pedido(conn, numero, produtos):
    """Salva os itens de um pedido da NuvemShop em pedido_itens (idempotente).
    Retorna True se os itens foram inseridos agora, False se já existiam."""
    existem = conn.execute(
        "SELECT COUNT(*) FROM pedido_itens WHERE pedido_numero = ?", (numero,)
    ).fetchone()[0]
    if existem:
        return False
    stmts = _build_item_stmts(numero, produtos)
    if stmts:
        conn.execute_batch(stmts)   # todos os itens numa requisição só
    return True


def _processar_estoque_pedidos():
    """Desconta do estoque os pedidos ainda não processados (estoque_processado=0).
    Seguro para rodar múltiplas vezes: cada pedido é marcado após processar."""
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        itens = conn.execute("""
            SELECT pi.pedido_numero, pi.nv_variant_id, pi.quantidade
            FROM pedido_itens pi
            JOIN pedidos p ON p.numero = pi.pedido_numero
            WHERE p.estoque_processado = 0
              AND pi.nv_variant_id IS NOT NULL
        """).fetchall()

        for item in itens:
            vid = item["nv_variant_id"]
            qty = item["quantidade"] or 1
            nr  = item["pedido_numero"]

            stock = conn.execute(
                "SELECT id, quantity, sku, produto_nome, variante_label FROM sku_stock WHERE nv_variant_id = ?", (vid,)
            ).fetchone()
            if not stock:
                continue   # variante ainda não tem estoque cadastrado → pula

            nova_qty = stock["quantity"] - qty
            conn.execute(
                "UPDATE sku_stock SET quantity=?, updated_at=? WHERE id=?",
                (nova_qty, agora, stock["id"]),
            )
            conn.execute(
                """INSERT INTO sku_stock_movements
                   (nv_variant_id, sku, tipo, quantidade, pedido_numero, created_at)
                   VALUES (?, ?, 'saida_venda', ?, ?, ?)""",
                (vid, stock["sku"], qty, nr, agora),
            )
            # Deduz da mesma quantidade em todos os produtos com mesmo nome + mesma SKU
            _deducao_sync_por_nome(
                conn, stock["produto_nome"], stock["variante_label"],
                qty, agora, skip_vid=vid, pedido_nr=nr, sku_origem=stock["sku"]
            )

        # Marca todos os pedidos não processados como processados
        conn.execute(
            "UPDATE pedidos SET estoque_processado = 1 WHERE estoque_processado = 0"
        )


def _sync_personalizacoes(conn):
    """Garante que todo pedido ativo com categoria='Personalizado' tenha entrada em personalizacoes."""
    pers_rows = conn.execute(
        "SELECT numero_pedido FROM personalizacoes"
    ).fetchall()
    ja_tem = {r["numero_pedido"] for r in pers_rows}

    novos = conn.execute(
        """SELECT numero, cliente FROM pedidos
           WHERE ativo = 1 AND categoria = 'Personalizado'"""
    ).fetchall()

    for p in novos:
        if p["numero"] not in ja_tem:
            conn.execute(
                """INSERT INTO personalizacoes
                   (numero_pedido, nome_personalizacao, status)
                   VALUES (?, ?, 'A SEPARAR')""",
                (p["numero"], p["cliente"] or None),
            )


# ── Routes — API ──────────────────────────────────────────────────────────────

@app.route("/api/pedidos", methods=["GET"])
@login_required
def get_pedidos():
    atualizar_status_db()
    hoje = date.today()
    with get_conn() as conn:
        rows = conn.execute(
            """SELECT id, numero, data_pedido, categoria, status, suspeito,
                      cliente, transportadora
               FROM pedidos WHERE ativo = 1 ORDER BY data_pedido ASC"""
        ).fetchall()
        # Map numero_pedido → latest personalization status (excluding ENVIADO)
        pers_rows = conn.execute(
            """SELECT numero_pedido, status FROM personalizacoes
               WHERE status != 'ENVIADO'"""
        ).fetchall()
    pers_map = {r["numero_pedido"]: r["status"] for r in pers_rows}

    result = []
    for r in rows:
        dp   = date.fromisoformat(r["data_pedido"])
        dias = calcular_dias_uteis(dp, hoje)
        result.append({
            "id":                   r["id"],
            "numero":               r["numero"],
            "data_pedido":          r["data_pedido"],
            "categoria":            r["categoria"],
            "status":               r["status"],
            "dias_uteis":           dias,
            "suspeito":             bool(r["suspeito"]),
            "cliente":              r["cliente"] or "",
            "transportadora":       r["transportadora"] or "",
            "personalizacao_status": pers_map.get(r["numero"]),
        })
    return jsonify(result)


@app.route("/api/pedidos", methods=["POST"])
@login_required
def add_pedido():
    data      = request.json or {}
    numero    = str(data.get("numero", "")).strip()
    data_str  = str(data.get("data_pedido", "")).strip()
    categoria = data.get("categoria", "Normal")

    if not numero or not data_str:
        return jsonify({"erro": "Número e data são obrigatórios"}), 400
    try:
        dp = date.fromisoformat(data_str)
    except ValueError:
        return jsonify({"erro": "Data inválida"}), 400

    if dp > date.today():
        return jsonify({"erro": "A data do pedido não pode ser no futuro"}), 400

    dias   = calcular_dias_uteis(dp, date.today())
    status = determinar_status(dias, categoria)

    with get_conn() as conn:
        existente = conn.execute(
            "SELECT id FROM pedidos WHERE numero = ?", (numero,)
        ).fetchone()
        if existente:
            return jsonify({"erro": f"Pedido #{numero} já existe na base de dados"}), 409
        conn.execute(
            "INSERT INTO pedidos (numero, data_pedido, categoria, status) VALUES (?, ?, ?, ?)",
            (numero, data_str, categoria, status),
        )
    return jsonify({"mensagem": "Pedido cadastrado"}), 201


@app.route("/api/pedidos/<int:pid>/enviar", methods=["POST"])
@login_required
def enviar_pedido(pid):
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        row = conn.execute(
            "SELECT status FROM pedidos WHERE id = ? AND ativo = 1", (pid,)
        ).fetchone()
        if not row:
            return jsonify({"erro": "Pedido não encontrado"}), 404
        conn.execute(
            "UPDATE pedidos SET ativo = 0, enviado_em = ?, status_ao_enviar = ? WHERE id = ?",
            (agora, row["status"], pid),
        )
    return jsonify({"mensagem": "Pedido marcado como enviado"})


@app.route("/api/pedidos/<int:pid>/cancelar", methods=["POST"])
@login_required
def cancelar_pedido(pid):
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        row = conn.execute(
            "SELECT status FROM pedidos WHERE id = ? AND ativo = 1", (pid,)
        ).fetchone()
        if not row:
            return jsonify({"erro": "Pedido não encontrado"}), 404
        conn.execute(
            "UPDATE pedidos SET ativo = 0, enviado_em = ?, status_ao_enviar = 'Cancelado' WHERE id = ?",
            (agora, pid),
        )
    return jsonify({"mensagem": "Pedido cancelado"})


@app.route("/api/pedidos/<int:pid>/categoria", methods=["POST"])
@login_required
def set_categoria(pid):
    nova = (request.json or {}).get("categoria", "Normal")
    if nova not in ("Normal", "Personalizado", "Internacional"):
        return jsonify({"erro": "Categoria inválida"}), 400
    with get_conn() as conn:
        row = conn.execute(
            "SELECT numero, cliente, data_pedido FROM pedidos WHERE id = ? AND ativo = 1", (pid,)
        ).fetchone()
        if not row:
            return jsonify({"erro": "Pedido não encontrado"}), 404
        dp   = date.fromisoformat(row["data_pedido"])
        dias = calcular_dias_uteis(dp, date.today())
        novo_status = determinar_status(dias, nova)
        conn.execute(
            "UPDATE pedidos SET categoria = ?, status = ? WHERE id = ?",
            (nova, novo_status, pid),
        )
        if nova == "Personalizado":
            # Cria entrada em personalizacoes se ainda não existir
            ja_existe = conn.execute(
                "SELECT id FROM personalizacoes WHERE numero_pedido = ?",
                (row["numero"],),
            ).fetchone()
            if not ja_existe:
                conn.execute(
                    """INSERT INTO personalizacoes
                       (numero_pedido, nome_personalizacao, status)
                       VALUES (?, ?, 'A SEPARAR')""",
                    (row["numero"], row["cliente"] or None),
                )
    return jsonify({"categoria": nova, "status": novo_status})


@app.route("/api/pedidos/<int:pid>/suspeito", methods=["POST"])
@login_required
def toggle_suspeito(pid):
    with get_conn() as conn:
        row = conn.execute(
            "SELECT suspeito FROM pedidos WHERE id = ? AND ativo = 1", (pid,)
        ).fetchone()
        if not row:
            return jsonify({"erro": "Pedido não encontrado"}), 404
        novo = 0 if row["suspeito"] else 1
        conn.execute("UPDATE pedidos SET suspeito = ? WHERE id = ?", (novo, pid))
    return jsonify({"suspeito": novo})


@app.route("/api/pedidos/<string:numero>/detalhes", methods=["GET"])
@login_required
def get_pedido_detalhes(numero):
    with get_conn() as conn:
        pedido = conn.execute(
            """SELECT numero, data_pedido, categoria, status, cliente, total,
                      pagamento, forma_pagamento, transportadora, suspeito,
                      ativo, enviado_em, status_ao_enviar, criado_em
               FROM pedidos WHERE numero = ?""",
            (numero,),
        ).fetchone()
        if not pedido:
            return jsonify({"erro": "Pedido não encontrado"}), 404
        itens = conn.execute(
            """SELECT produto_nome, variante, quantidade, preco_unit, imagem_url
               FROM pedido_itens WHERE pedido_numero = ? ORDER BY id""",
            (numero,),
        ).fetchall()
        pers = conn.execute(
            """SELECT nome_personalizacao, numero_personalizacao, status, observacao
               FROM personalizacoes WHERE numero_pedido = ? ORDER BY id DESC LIMIT 1""",
            (numero,),
        ).fetchone()
    return jsonify({
        "pedido": dict(pedido),
        "itens":  [dict(i) for i in itens],
        "personalizacao": dict(pers) if pers else None,
    })


@app.route("/api/pedidos/<int:pid>", methods=["DELETE"])
@login_required
def delete_pedido(pid):
    with get_conn() as conn:
        conn.execute("DELETE FROM pedidos WHERE id = ?", (pid,))
    return jsonify({"mensagem": "Pedido excluído"})


@app.route("/api/stats")
@login_required
def get_stats():
    atualizar_status_db()
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT status, COUNT(*) as total FROM pedidos WHERE ativo = 1 AND (suspeito IS NULL OR suspeito = 0) GROUP BY status"
        ).fetchall()
        suspeitos = conn.execute(
            "SELECT COUNT(*) as total FROM pedidos WHERE ativo = 1 AND suspeito = 1"
        ).fetchone()["total"]
        enviados = conn.execute(
            "SELECT COUNT(*) as total FROM pedidos WHERE ativo = 0"
        ).fetchone()["total"]
    counts = {s: 0 for s in STATUS_ORDER}
    for r in rows:
        if r["status"] in counts:
            counts[r["status"]] = r["total"]
    total = sum(counts.values())
    return jsonify({
        "total": total,
        "suspeitos": suspeitos,
        "enviados": enviados,
        "por_status": [
            {
                "status":     s,
                "total":      counts[s],
                "percentual": round(counts[s] / total * 100, 1) if total else 0,
            }
            for s in STATUS_ORDER
        ],
    })


@app.route("/api/atualizar", methods=["POST"])
@login_required
def atualizar_manual():
    atualizar_status_db()
    return jsonify({"mensagem": "Status atualizados"})


@app.route("/api/relatorios")
@login_required
def get_relatorios():
    inicio_param = request.args.get("inicio", "")
    fim_param    = request.args.get("fim", "")
    if inicio_param and fim_param:
        inicio = inicio_param
        fim    = fim_param
    else:
        periodo = request.args.get("periodo", "30dias")
        inicio  = periodo_para_data_inicio(periodo).isoformat()
        fim     = date.today().isoformat()

    with get_conn() as conn:
        rows = conn.execute(
            """SELECT numero, data_pedido, categoria, status_ao_enviar, enviado_em,
                      cliente, transportadora
               FROM pedidos
               WHERE ativo = 0
                 AND date(enviado_em) >= ?
                 AND date(enviado_em) <= ?
               ORDER BY enviado_em DESC""",
            (inicio, fim),
        ).fetchall()

    pedidos    = [dict(r) for r in rows]
    total      = len(pedidos)
    cancelados = sum(1 for p in pedidos if p.get("status_ao_enviar") == "Cancelado")
    enviados   = total - cancelados
    counts     = {s: 0 for s in STATUS_ORDER}
    for p in pedidos:
        s = p.get("status_ao_enviar") or ""
        if s in counts:
            counts[s] += 1

    carrier_counts: dict = {}
    for p in pedidos:
        if p.get("status_ao_enviar") == "Cancelado":
            continue
        t = p.get("transportadora") or "Sem transportadora"
        carrier_counts[t] = carrier_counts.get(t, 0) + 1
    por_transportadora = sorted(
        [{"transportadora": t, "total": n} for t, n in carrier_counts.items()],
        key=lambda x: -x["total"],
    )

    return jsonify({
        "total":              total,
        "enviados":           enviados,
        "cancelados":         cancelados,
        "por_status": [
            {
                "status":     s,
                "total":      counts[s],
                "percentual": round(counts[s] / enviados * 100, 1) if enviados else 0,
            }
            for s in STATUS_ORDER
        ],
        "por_transportadora": por_transportadora,
        "pedidos":            pedidos,
    })


@app.route("/api/exportar")
@login_required
def exportar():
    inicio_param = request.args.get("inicio", "")
    fim_param    = request.args.get("fim", "")
    if inicio_param and fim_param:
        inicio = inicio_param
        fim    = fim_param
    else:
        periodo = request.args.get("periodo", "30dias")
        inicio  = periodo_para_data_inicio(periodo).isoformat()
        fim     = date.today().isoformat()

    with get_conn() as conn:
        rows = conn.execute(
            """SELECT numero, data_pedido, categoria, status_ao_enviar, enviado_em
               FROM pedidos
               WHERE ativo = 0
                 AND date(enviado_em) >= ?
                 AND date(enviado_em) <= ?
               ORDER BY enviado_em DESC""",
            (inicio, fim),
        ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório de Pedidos"
    ws.append(["Número", "Cliente", "Data do Pedido", "Categoria", "Transportadora", "Status ao Enviar", "Data de Envio"])
    for r in rows:
        ws.append([r["numero"], r["cliente"], r["data_pedido"], r["categoria"],
                   r["transportadora"], r["status_ao_enviar"], r["enviado_em"]])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        download_name=f"relatorio_{periodo}.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── Import helpers ────────────────────────────────────────────────────────────

MESES_PT = {"jan":1,"fev":2,"mar":3,"abr":4,"mai":5,"jun":6,
             "jul":7,"ago":8,"set":9,"out":10,"nov":11,"dez":12}

def parse_date(s: str):
    s = s.strip()
    try:
        return date.fromisoformat(s[:10]).isoformat()
    except Exception:
        pass
    m = re.match(r"(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100: y += 2000
        try: return date(y, mo, d).isoformat()
        except Exception: pass
    m = re.match(r"(\d{1,2})\s+([a-záéíóú]{3})", s.lower())
    if m:
        mes = MESES_PT.get(m.group(2)[:3])
        if mes:
            try: return date(date.today().year, mes, int(m.group(1))).isoformat()
            except Exception: pass
    return None


@app.route("/api/importar/imagem", methods=["POST"])
@login_required
def importar_imagem():
    try:
        import anthropic as ant
    except ImportError:
        return jsonify({"erro": "Pacote anthropic não instalado."}), 500

    api_key = os.getenv("ANTHROPIC_API_KEY", "")
    if not api_key or "cole" in api_key:
        return jsonify({"erro": "ANTHROPIC_API_KEY não configurada no arquivo .env"}), 400

    data       = request.json or {}
    imagem_b64 = data.get("imagem", "")
    media_type = data.get("tipo", "image/png")
    categoria  = data.get("categoria", "Normal")

    if not imagem_b64:
        return jsonify({"erro": "Nenhuma imagem recebida"}), 400

    ano    = date.today().year
    prompt = (
        f"Esta é uma captura de tela do painel de pedidos da NuvemShop.\n"
        f"Extraia TODOS os pedidos visíveis com os campos abaixo.\n"
        f"Retorne APENAS JSON válido, sem texto adicional:\n"
        f'{{"pedidos":['
        f'{{"numero":"6475","data":"{ano}-05-06","cliente":"Diego Sacramento",'
        f'"total":"339.16","pagamento":"Recebido","envio":"Enviada","transportadora":"Correios PAC"}}'
        f']}}\n\n'
        f"Regras:\n"
        f"- numero: somente dígitos, sem #\n"
        f"- data: YYYY-MM-DD, ano={ano}. Meses: jan=01 fev=02 mar=03 abr=04 mai=05 jun=06 jul=07 ago=08 set=09 out=10 nov=11 dez=12\n"
        f"- cliente: nome completo do cliente\n"
        f"- total: valor numérico sem R$ (ex: '339.16')\n"
        f"- pagamento: 'Recebido' se mostrar 'Recebido' ou 'Pago'; 'Pendente' se não pago ou pendente\n"
        f"- envio: use EXATAMENTE uma destas opções:\n"
        f"    'Enviada'   → se mostrar 'Enviada' (pacote despachado)\n"
        f"    'Entregue'  → se mostrar 'Entregue' (entregue ao cliente)\n"
        f"    'Por enviar'→ se ainda não foi enviado\n"
        f"- transportadora: ex 'Correios PAC', 'Correios SEDEX', 'Loggi Pontos', 'Nuvem Envio - Correios PAC'\n"
        f"- ATENÇÃO: se a tela mostrar todos os pedidos como Enviada ou Entregue, retorne isso para TODOS\n"
        f"- Inclua TODOS os pedidos da imagem sem exceção"
    )

    prompt2 = (
        f"Esta é a mesma imagem do painel da NuvemShop.\n"
        f"Analise EXCLUSIVAMENTE o status de envio de CADA pedido visível.\n"
        f"Procure com atenção badges coloridos, ícones e textos como 'Enviada', 'Entregue', 'Por enviar'.\n"
        f"Retorne APENAS JSON válido, sem texto adicional:\n"
        f'{{"pedidos":[{{"numero":"6475","envio":"Enviada"}}]}}\n\n'
        f"Regras:\n"
        f"- numero: somente dígitos, sem #\n"
        f"- envio: use EXATAMENTE uma das opções:\n"
        f"    'Enviada'    → badge/texto 'Enviada' (pacote despachado)\n"
        f"    'Entregue'   → badge/texto 'Entregue' (entregue ao cliente)\n"
        f"    'Por enviar' → ainda não despachado\n"
        f"- Inclua TODOS os pedidos da imagem"
    )

    try:
        client = ant.Anthropic(api_key=api_key)

        # Primeira análise: extração completa
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=2048,
            messages=[{"role": "user", "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": imagem_b64}},
                {"type": "text",  "text": prompt},
            ]}],
        )

        # Segunda análise: foco exclusivo no status de envio
        resp2 = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1024,
            messages=[{"role": "user", "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": imagem_b64}},
                {"type": "text",  "text": prompt2},
            ]}],
        )

    except Exception as e:
        msg = str(e)
        if "credit" in msg.lower() or "balance" in msg.lower() or "quota" in msg.lower() or "billing" in msg.lower():
            return jsonify({"erro": "Saldo insuficiente na conta Anthropic. Adicione crédito em console.anthropic.com → Faturamento."}), 402
        return jsonify({"erro": f"Erro na API: {msg[:300]}"}), 500

    # Parse primeira análise
    texto = resp.content[0].text.strip()
    texto = re.sub(r"```(?:json)?", "", texto).strip().rstrip("`").strip()
    try:
        pedidos = json.loads(texto).get("pedidos", [])
    except Exception:
        return jsonify({"erro": f"Não foi possível interpretar a resposta: {texto[:200]}"}), 400

    # Parse segunda análise e faz merge do campo envio
    texto2 = resp2.content[0].text.strip()
    texto2 = re.sub(r"```(?:json)?", "", texto2).strip().rstrip("`").strip()
    try:
        pedidos2   = json.loads(texto2).get("pedidos", [])
        envio_map2 = {str(p2.get("numero", "")).strip(): p2.get("envio", "") for p2 in pedidos2}
        for p in pedidos:
            num = str(p.get("numero", "")).strip()
            if num in envio_map2 and envio_map2[num]:
                p["envio"] = envio_map2[num]
    except Exception:
        pass  # segunda análise falhou: mantém resultado da primeira

    for p in pedidos:
        p["categoria"] = categoria

    return jsonify({"pedidos": pedidos, "total": len(pedidos)})


@app.route("/api/importar/arquivo", methods=["POST"])
@login_required
def importar_arquivo():
    arquivo = request.files.get("arquivo")
    if not arquivo:
        return jsonify({"erro": "Nenhum arquivo enviado"}), 400

    categoria = request.form.get("categoria", "Normal")
    nome      = arquivo.filename.lower()
    pedidos   = []

    try:
        if nome.endswith(".csv"):
            conteudo = arquivo.read().decode("utf-8-sig")
            leitor   = csv.DictReader(StringIO(conteudo))
            for linha in leitor:
                numero = data_str = None
                for chave, val in linha.items():
                    ch = chave.lower()
                    if not numero and any(x in ch for x in ["número","numero","pedido","order","nº","#"]):
                        numero = str(val or "").strip().lstrip("#").strip()
                    if not data_str and any(x in ch for x in ["data","date","criação","criacao"]):
                        data_str = str(val or "").strip()
                if numero and data_str:
                    d = parse_date(data_str)
                    if d:
                        pedidos.append({"numero": numero, "data": d, "categoria": categoria})

        elif nome.endswith((".xlsx", ".xls")):
            wb      = openpyxl.load_workbook(BytesIO(arquivo.read()), data_only=True)
            ws      = wb.active
            headers = [str(c.value or "").lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            num_col  = next((i for i, h in enumerate(headers) if any(x in h for x in ["número","numero","pedido","nº"])), None)
            date_col = next((i for i, h in enumerate(headers) if any(x in h for x in ["data","date","criação","criacao"])), None)
            if num_col is None or date_col is None:
                return jsonify({"erro": "Colunas de número e data não encontradas. Verifique o arquivo."}), 400
            for row in ws.iter_rows(min_row=2, values_only=True):
                numero = str(row[num_col] or "").strip().lstrip("#").strip()
                val    = row[date_col]
                if isinstance(val, datetime):
                    d = val.date().isoformat()
                elif isinstance(val, date):
                    d = val.isoformat()
                else:
                    d = parse_date(str(val or ""))
                if numero and d:
                    pedidos.append({"numero": numero, "data": d, "categoria": categoria})
        else:
            return jsonify({"erro": "Formato não suportado. Use .csv ou .xlsx"}), 400

    except Exception as e:
        return jsonify({"erro": f"Erro ao processar arquivo: {e}"}), 400

    return jsonify({"pedidos": pedidos, "total": len(pedidos)})


@app.route("/api/importar/verificar", methods=["POST"])
@login_required
def verificar_duplicados():
    numeros = (request.json or {}).get("numeros", [])
    if not numeros:
        return jsonify({"duplicados": []})
    with get_conn() as conn:
        placeholders = ",".join("?" * len(numeros))
        rows = conn.execute(
            f"SELECT numero FROM pedidos WHERE numero IN ({placeholders})",
            numeros,
        ).fetchall()
    return jsonify({"duplicados": [r["numero"] for r in rows]})


@app.route("/api/importar/confirmar", methods=["POST"])
@login_required
def importar_confirmar():
    data    = request.json or {}
    pedidos = data.get("pedidos", [])
    hoje    = date.today()
    salvos = ignorados = ja_enviados = 0

    duplicados = 0
    with get_conn() as conn:
        for p in pedidos:
            numero         = str(p.get("numero", "")).strip()
            data_str       = p.get("data", "")
            categoria      = p.get("categoria", "Normal")
            pagamento      = str(p.get("pagamento", "")).strip()
            envio          = str(p.get("envio", "")).strip()
            cliente        = p.get("cliente", "")
            total          = p.get("total", "")
            transportadora = p.get("transportadora", "")

            if not numero or not data_str:
                continue

            # Ignora duplicatas — mesmo número já existe na base
            existente = conn.execute(
                "SELECT id FROM pedidos WHERE numero = ?", (numero,)
            ).fetchone()
            if existente:
                duplicados += 1
                continue

            # Ignora pedidos com pagamento pendente
            if pagamento and "recebido" not in pagamento.lower() and "pago" not in pagamento.lower():
                ignorados += 1
                continue

            try:
                dp = date.fromisoformat(data_str)
            except ValueError:
                continue

            if dp > hoje:
                dp = hoje  # data futura: trata como hoje

            dias   = calcular_dias_uteis(dp, hoje)
            status = determinar_status(dias, categoria)

            is_enviada = envio and any(s in envio.lower() for s in ["enviada", "entregue"])

            try:
                if is_enviada:
                    conn.execute(
                        """INSERT INTO pedidos
                           (numero, data_pedido, categoria, status, cliente, total, pagamento, transportadora,
                            ativo, enviado_em, status_ao_enviar)
                           VALUES (?,?,?,?,?,?,?,?,0,?,?)""",
                        (numero, data_str, categoria, status, cliente, total, pagamento, transportadora,
                         hoje.isoformat(), status),
                    )
                    ja_enviados += 1
                else:
                    conn.execute(
                        """INSERT INTO pedidos
                           (numero, data_pedido, categoria, status, cliente, total, pagamento, transportadora)
                           VALUES (?,?,?,?,?,?,?,?)""",
                        (numero, data_str, categoria, status, cliente, total, pagamento, transportadora),
                    )
                salvos += 1
            except Exception:
                pass

    with get_conn() as conn:
        _sync_personalizacoes(conn)

    partes = [f"{salvos} pedido(s) importado(s)"]
    if duplicados:
        partes.append(f"{duplicados} duplicado(s) ignorado(s)")
    if ignorados:
        partes.append(f"{ignorados} ignorado(s) por pagamento pendente")
    if ja_enviados:
        partes.append(f"{ja_enviados} já enviado(s) foram direto para o relatório")

    return jsonify({"mensagem": " — ".join(partes), "salvos": salvos,
                    "duplicados": duplicados, "ignorados": ignorados, "ja_enviados": ja_enviados})


# ── Notificações ─────────────────────────────────────────────────────────────

@app.route("/api/notificacoes")
@login_required
def get_notificacoes():
    hoje = date.today()

    # Próximo dia útil
    feriados = holidays.Brazil(years=[hoje.year, hoje.year + 1])
    proximo_du = hoje + timedelta(days=1)
    while proximo_du.weekday() >= 5 or proximo_du in feriados:
        proximo_du += timedelta(days=1)

    with get_conn() as conn:
        rows = conn.execute(
            "SELECT id, numero, data_pedido, categoria, status FROM pedidos WHERE ativo = 1"
        ).fetchall()

    alertas = []
    for r in rows:
        dp           = date.fromisoformat(r["data_pedido"])
        status_hoje  = determinar_status(calcular_dias_uteis(dp, hoje),      r["categoria"])
        status_amanha= determinar_status(calcular_dias_uteis(dp, proximo_du), r["categoria"])
        if status_amanha != status_hoje:
            alertas.append({
                "id":           r["id"],
                "numero":       r["numero"],
                "status_atual": status_hoje,
                "status_novo":  status_amanha,
            })

    STATUS_PRIO = ["Atraso crítico", "Atraso moderado", "Atraso leve"]
    grupos: dict = {}
    for a in alertas:
        s = a["status_novo"]
        grupos.setdefault(s, []).append(a["numero"])

    return jsonify({
        "total": len(alertas),
        "grupos": [
            {"status": s, "total": len(grupos[s]), "numeros": grupos[s]}
            for s in STATUS_PRIO if s in grupos
        ],
    })


# ── Personalizações ──────────────────────────────────────────────────────────

@app.route("/api/personalizacoes/stats")
@login_required
def pers_stats():
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT status, COUNT(*) as n FROM personalizacoes GROUP BY status"
        ).fetchall()
    counts = {s: 0 for s in PERS_STATUS}
    for r in rows:
        if r["status"] in counts:
            counts[r["status"]] = r["n"]
    total_ativas = sum(counts[s] for s in PERS_STATUS if s != "ENVIADO")
    return jsonify({"por_status": counts, "total_ativas": total_ativas})


@app.route("/api/personalizacoes", methods=["GET"])
@login_required
def get_personalizacoes():
    with get_conn() as conn:
        _sync_personalizacoes(conn)
        rows = conn.execute(
            """SELECT id, numero_pedido, nome_personalizacao, numero_personalizacao,
                      status, observacao, criado_em, atualizado_em
               FROM personalizacoes ORDER BY
               CASE status
                   WHEN 'A SEPARAR'         THEN 1
                   WHEN 'SEPARAÇÃO'         THEN 2
                   WHEN 'NA PERSONALIZAÇÃO' THEN 3
                   WHEN 'PRONTO'            THEN 4
                   WHEN 'ENVIADO'           THEN 5
                   ELSE 6
               END, id ASC"""
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/personalizacoes", methods=["POST"])
@login_required
def criar_personalizacao():
    data   = request.json or {}
    numero = str(data.get("numero_pedido", "")).strip()
    if not numero:
        return jsonify({"erro": "Número do pedido é obrigatório"}), 400
    nome     = str(data.get("nome_personalizacao", "")).strip()
    num_pers = str(data.get("numero_personalizacao", "")).strip()
    obs      = str(data.get("observacao", "")).strip()
    with get_conn() as conn:
        cur = conn.execute(
            """INSERT INTO personalizacoes
               (numero_pedido, nome_personalizacao, numero_personalizacao, observacao)
               VALUES (?, ?, ?, ?)""",
            (numero, nome or None, num_pers or None, obs or None),
        )
    return jsonify({"mensagem": "Personalização criada", "id": cur.lastrowid}), 201


@app.route("/api/personalizacoes/<int:pid>/status", methods=["POST"])
@login_required
def atualizar_pers_status(pid):
    novo = (request.json or {}).get("status", "")
    if novo not in PERS_STATUS:
        return jsonify({"erro": "Status inválido"}), 400
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        row = conn.execute(
            "SELECT id FROM personalizacoes WHERE id = ?", (pid,)
        ).fetchone()
        if not row:
            return jsonify({"erro": "Personalização não encontrada"}), 404
        conn.execute(
            "UPDATE personalizacoes SET status = ?, atualizado_em = ? WHERE id = ?",
            (novo, agora, pid),
        )
    return jsonify({"status": novo})


@app.route("/api/personalizacoes/<int:pid>", methods=["DELETE"])
@login_required
def deletar_personalizacao(pid):
    with get_conn() as conn:
        conn.execute("DELETE FROM personalizacoes WHERE id = ?", (pid,))
    return jsonify({"mensagem": "Personalização excluída"})


@app.route("/api/personalizacoes/<int:pid>", methods=["PATCH"])
@login_required
def editar_personalizacao(pid):
    data = request.json or {}
    with get_conn() as conn:
        row = conn.execute(
            "SELECT id FROM personalizacoes WHERE id = ?", (pid,)
        ).fetchone()
        if not row:
            return jsonify({"erro": "Personalização não encontrada"}), 404
        agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn.execute(
            """UPDATE personalizacoes
               SET nome_personalizacao   = ?,
                   numero_personalizacao = ?,
                   observacao            = ?,
                   atualizado_em         = ?
               WHERE id = ?""",
            (
                data.get("nome_personalizacao") or None,
                data.get("numero_personalizacao") or None,
                data.get("observacao") or None,
                agora, pid,
            ),
        )
    return jsonify({"mensagem": "Atualizado"})


# ── Manutenção ───────────────────────────────────────────────────────────────

@app.route("/api/manutencao/corrigir-reconciliacao", methods=["POST"])
@login_required
def corrigir_reconciliacao():
    """Corrige pedidos marcados erroneamente pela reconciliação automática.
    Aplica: status_ao_enviar='NO PRAZO', enviado_em=data_pedido+3 dias.
    Filtro seguro: apenas pedidos sem romaneio, enviados hoje, com status de atraso.
    """
    hoje = date.today().isoformat()
    with get_conn() as conn:
        # Conta quantos serão afetados antes de alterar
        antes = conn.execute(
            """SELECT COUNT(*) as n FROM pedidos
               WHERE ativo = 0
                 AND romaneio_id IS NULL
                 AND date(enviado_em) = ?
                 AND status_ao_enviar IN ('Atraso crítico', 'Atraso moderado', 'Atraso leve')""",
            (hoje,),
        ).fetchone()["n"]

        if antes == 0:
            return jsonify({"corrigidos": 0, "mensagem": "Nenhum pedido para corrigir."})

        conn.execute(
            """UPDATE pedidos
               SET status_ao_enviar = 'NO PRAZO',
                   enviado_em = CASE
                       WHEN date(data_pedido, '+3 days') <= ?
                       THEN date(data_pedido, '+3 days')
                       ELSE ?
                   END
               WHERE ativo = 0
                 AND romaneio_id IS NULL
                 AND date(enviado_em) = ?
                 AND status_ao_enviar IN ('Atraso crítico', 'Atraso moderado', 'Atraso leve')""",
            (hoje, hoje, hoje),
        )
    return jsonify({
        "corrigidos": antes,
        "mensagem": f"{antes} pedidos corrigidos: status → NO PRAZO, data → 3 dias após compra",
    })


@app.route("/api/manutencao/pendentes-correcao")
@login_required
def pendentes_correcao():
    hoje = date.today().isoformat()
    with get_conn() as conn:
        n = conn.execute(
            """SELECT COUNT(*) as n FROM pedidos
               WHERE ativo = 0
                 AND romaneio_id IS NULL
                 AND date(enviado_em) = ?
                 AND status_ao_enviar IN ('Atraso crítico', 'Atraso moderado', 'Atraso leve')""",
            (hoje,),
        ).fetchone()["n"]
    return jsonify({"pendentes": n})


# ── Romaneio routes ──────────────────────────────────────────────────────────

@app.route("/api/romaneios", methods=["GET"])
@login_required
def get_romaneios():
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT id, criado_em, total, observacao FROM romaneios ORDER BY id DESC"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/romaneios", methods=["POST"])
@login_required
def criar_romaneio():
    data       = request.json or {}
    ids        = data.get("ids", [])
    observacao = data.get("observacao", "").strip()

    if not ids:
        return jsonify({"erro": "Nenhum pedido selecionado"}), 400

    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with get_conn() as conn:
        cur = conn.execute(
            "INSERT INTO romaneios (total, observacao) VALUES (?, ?)",
            (len(ids), observacao or None),
        )
        romaneio_id = cur.lastrowid

        salvos = 0
        for pid in ids:
            row = conn.execute(
                "SELECT status FROM pedidos WHERE id = ? AND ativo = 1", (pid,)
            ).fetchone()
            if row:
                conn.execute(
                    """UPDATE pedidos
                       SET ativo = 0, enviado_em = ?, status_ao_enviar = ?, romaneio_id = ?
                       WHERE id = ?""",
                    (agora, row["status"], romaneio_id, pid),
                )
                salvos += 1

        if salvos == 0:
            conn.execute("DELETE FROM romaneios WHERE id = ?", (romaneio_id,))
            return jsonify({"erro": "Nenhum pedido pôde ser enviado. Os pedidos selecionados já foram marcados como enviados (pela sincronização automática)?"}), 400

        conn.execute("UPDATE romaneios SET total = ? WHERE id = ?", (salvos, romaneio_id))

    return jsonify({
        "mensagem": f"Romaneio #{romaneio_id} criado com {salvos} pedido(s)",
        "id":       romaneio_id,
        "total":    salvos,
    })


@app.route("/api/romaneios/<int:rid>", methods=["GET"])
@login_required
def get_romaneio(rid):
    with get_conn() as conn:
        rom = conn.execute(
            "SELECT id, criado_em, total, observacao FROM romaneios WHERE id = ?", (rid,)
        ).fetchone()
        if not rom:
            return jsonify({"erro": "Romaneio não encontrado"}), 404
        pedidos = conn.execute(
            """SELECT numero, data_pedido, cliente, transportadora, status_ao_enviar, enviado_em
               FROM pedidos WHERE romaneio_id = ? ORDER BY numero""",
            (rid,),
        ).fetchall()

    pedidos_list = [dict(p) for p in pedidos]
    grupos: dict = {}
    for p in pedidos_list:
        t = p.get("transportadora") or "Sem transportadora"
        grupos.setdefault(t, []).append(p)
    por_transportadora = [{"transportadora": t, "pedidos": ps} for t, ps in grupos.items()]

    return jsonify({
        "id":               rom["id"],
        "criado_em":        rom["criado_em"],
        "total":            rom["total"],
        "observacao":       rom["observacao"],
        "pedidos":          pedidos_list,
        "por_transportadora": por_transportadora,
    })


@app.route("/api/romaneios/<int:rid>/exportar", methods=["GET"])
@login_required
def exportar_romaneio(rid):
    with get_conn() as conn:
        rom = conn.execute(
            "SELECT id, criado_em, total, observacao FROM romaneios WHERE id = ?", (rid,)
        ).fetchone()
        if not rom:
            return jsonify({"erro": "Romaneio não encontrado"}), 404
        pedidos = conn.execute(
            """SELECT numero, data_pedido, cliente, transportadora, status_ao_enviar, enviado_em
               FROM pedidos WHERE romaneio_id = ? ORDER BY numero""",
            (rid,),
        ).fetchall()

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    pedidos_list = [dict(p) for p in pedidos]
    grupos: dict = {}
    for p in pedidos_list:
        t = p.get("transportadora") or "Sem transportadora"
        grupos.setdefault(t, []).append(p)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Romaneio {rid}"

    # Cabeçalho geral
    ws.append([f"Romaneio #{rid}"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([f"Gerado em: {rom['criado_em']}"])
    if rom["observacao"]:
        ws.append([f"Obs: {rom['observacao']}"])
    ws.append([f"Total de pedidos: {rom['total']}"])
    ws.append([])

    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1F6FEB")
    hdr_cols  = ["Número", "Cliente", "Data do Pedido", "Status ao Enviar", "Data de Envio"]

    for transportadora, itens in grupos.items():
        # Linha de seção por transportadora
        row_grp = ws.max_row + 1
        ws.append([f"{transportadora.upper()}  —  {len(itens)} pedido(s)"])
        cell_grp = ws.cell(row=row_grp, column=1)
        cell_grp.font  = Font(bold=True, size=11, color="FFFFFF")
        cell_grp.fill  = PatternFill("solid", fgColor="21262D")
        cell_grp.alignment = Alignment(vertical="center")
        ws.merge_cells(start_row=row_grp, start_column=1, end_row=row_grp, end_column=5)

        # Cabeçalho das colunas
        ws.append(hdr_cols)
        hdr_row = ws.max_row
        for col in range(1, 6):
            c = ws.cell(row=hdr_row, column=col)
            c.font = hdr_font
            c.fill = hdr_fill

        for p in itens:
            ws.append([
                p["numero"], p["cliente"] or "", p["data_pedido"],
                p["status_ao_enviar"] or "", (p["enviado_em"] or "")[:10],
            ])

        ws.append([])

    # Ajusta largura das colunas
    for col_idx in range(1, 6):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = [15, 28, 15, 18, 15][col_idx - 1]

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        download_name=f"romaneio_{rid}.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── NuvemShop sync ───────────────────────────────────────────────────────────

def _get_nv_credentials():
    """Lê store_id e access_token: primeiro das env vars, depois do banco."""
    store_id = os.getenv("NUVEMSHOP_STORE_ID", "")
    token    = os.getenv("NUVEMSHOP_ACCESS_TOKEN", "")
    if not store_id or not token:
        try:
            with get_conn() as conn:
                row_id = conn.execute(
                    "SELECT valor FROM config WHERE chave='nuvemshop_store_id'"
                ).fetchone()
                row_tk = conn.execute(
                    "SELECT valor FROM config WHERE chave='nuvemshop_access_token'"
                ).fetchone()
                if row_id and row_id["valor"]:
                    store_id = row_id["valor"]
                if row_tk and row_tk["valor"]:
                    token = row_tk["valor"]
        except Exception:
            pass
    return store_id, token


def _nuvemshop_headers():
    _, token = _get_nv_credentials()
    return {
        "Authentication": f"bearer {token}",
        "User-Agent": "GS Mantos Interno (contato@gsmantos.com.br)",
        "Content-Type": "application/json",
    }

def _parse_order_fields(o, hoje=None):
    """Extrai e normaliza os campos de um pedido vindo da API NuvemShop.
    Retorna um dict pronto para inserção, ou None se o pedido for inválido/cancelado."""
    if hoje is None:
        hoje = date.today()

    # Ignora pedidos cancelados
    if o.get("status") == "cancelled" or o.get("payment_status") in ("voided", "refunded"):
        return None

    numero   = str(o.get("number", "")).strip()
    data_str = o.get("created_at", "")[:10]
    if not numero or not data_str:
        return None

    cliente = o.get("contact_name", "") or ""
    total   = str(o.get("total", ""))

    shipping_status = o.get("shipping_status", "")
    if shipping_status == "shipped":
        envio = "Enviada"
    elif shipping_status == "delivered":
        envio = "Entregue"
    else:
        envio = "Por enviar"

    transportadora = ""
    for f in (o.get("fulfillments") or []):
        carrier = (f.get("shipping") or {}).get("carrier", {}).get("name", "")
        option  = (f.get("shipping") or {}).get("option",  {}).get("name", "")
        if carrier or option:
            transportadora = f"{carrier} - {option}".strip(" -")
        break
    if not transportadora:
        transportadora = o.get("shipping_option", "") or ""
    transportadora = transportadora.replace("Nuvem Envio - ", "")

    gateway = o.get("gateway_name", "") or ""
    method  = (o.get("payment_details") or {}).get("method", "") or ""
    if gateway and method:
        forma_pagamento = f"{gateway} - {method}"
    elif gateway:
        forma_pagamento = gateway
    else:
        forma_pagamento = method or ""

    try:
        dp = date.fromisoformat(data_str)
    except ValueError:
        return None
    if dp > hoje:
        dp = hoje

    dias   = calcular_dias_uteis(dp, hoje)
    status = determinar_status(dias, "Normal")

    return {
        "numero":          numero,
        "data_str":        data_str,
        "cliente":         cliente,
        "total":           total,
        "envio":           envio,
        "transportadora":  transportadora,
        "forma_pagamento": forma_pagamento,
        "status":          status,
        "is_env":          envio in ("Enviada", "Entregue"),
        "produtos":        o.get("products") or [],
    }


def _processar_order(conn, o):
    """Processa UM pedido da NuvemShop e insere no banco se novo (caminho usado pelo webhook).
    Retorna 'salvo'|'ja_enviado'|'duplicado'|'skip'."""
    c = _parse_order_fields(o)
    if not c:
        return "skip"

    numero, produtos = c["numero"], c["produtos"]

    existente = conn.execute(
        "SELECT id FROM pedidos WHERE numero = ?", (numero,)
    ).fetchone()
    if existente:
        if produtos:
            _salvar_itens_pedido(conn, numero, produtos)
        return "duplicado"

    hoje = date.today()
    try:
        if c["is_env"]:
            conn.execute(
                """INSERT INTO pedidos
                   (numero,data_pedido,categoria,status,cliente,total,pagamento,forma_pagamento,
                    transportadora,ativo,enviado_em,status_ao_enviar)
                   VALUES (?,?,?,?,?,?,?,?,?,0,?,?)""",
                (numero, c["data_str"], "Normal", c["status"], c["cliente"], c["total"],
                 "Recebido", c["forma_pagamento"], c["transportadora"],
                 hoje.isoformat(), c["status"]),
            )
            if produtos:
                _salvar_itens_pedido(conn, numero, produtos)
            return "ja_enviado"
        else:
            conn.execute(
                """INSERT INTO pedidos
                   (numero,data_pedido,categoria,status,cliente,total,pagamento,forma_pagamento,transportadora)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                (numero, c["data_str"], "Normal", c["status"], c["cliente"], c["total"],
                 "Recebido", c["forma_pagamento"], c["transportadora"]),
            )
            if produtos:
                _salvar_itens_pedido(conn, numero, produtos)
            return "salvo"
    except Exception:
        return "skip"


def sincronizar_nuvemshop():
    import urllib.request as ureq
    store_id, token = _get_nv_credentials()
    if not store_id or not token:
        return {"erro": "NuvemShop não configurada"}

    headers = _nuvemshop_headers()

    # ── Pré-carrega o estado do banco numa única leitura ────────────────────────
    # Em vez de 1 SELECT por pedido (centenas de viagens de rede), carrega de uma vez
    # o conjunto de números já cadastrados e quais já têm itens salvos.
    with get_conn() as conn:
        row = conn.execute("SELECT valor FROM config WHERE chave='ultima_sincronizacao'").fetchone()
        ultima_sync = row["valor"] if row else None
        existentes = {r["numero"] for r in
                      conn.execute("SELECT numero FROM pedidos").fetchall()}
        com_itens  = {r["pedido_numero"] for r in
                      conn.execute("SELECT DISTINCT pedido_numero FROM pedido_itens").fetchall()}

    salvos = duplicados = ja_enviados = 0
    hoje = date.today()

    # Instrumentação: mede o tempo gasto em cada fase (retornado no resultado)
    import time
    _ult = [time.time()]
    tempos = {}
    def _marco(nome):
        agora_t = time.time()
        tempos[nome] = round(agora_t - _ult[0], 2)
        _ult[0] = agora_t

    # Filtro incremental: a API NuvemShop usa ISO 8601 (NÃO unix timestamp — o
    # código antigo passava timestamp unix, que a API ignorava → baixava TUDO sempre).
    # updated_at_min pega qualquer pedido que mudou de status desde a última sync.
    # Recua 1 dia como margem de segurança contra diferenças de fuso horário.
    iso_q = ""
    if ultima_sync:
        try:
            from datetime import timedelta as _td
            dt = datetime.fromisoformat(ultima_sync) - _td(days=1)
            iso_q = dt.strftime("%Y-%m-%d")   # data simples ISO 8601 (sem hora/timezone)
        except Exception:
            iso_q = ""

    # ── Acumulador: todas as escritas vão para 'pending' e são gravadas em lotes ──
    pending = []

    def _flush():
        """Grava os statements acumulados em lotes de 100 (1 requisição por lote)."""
        if not pending:
            return
        with get_conn() as conn:
            for i in range(0, len(pending), 100):
                lote = pending[i:i + 100]
                try:
                    conn.execute_batch(lote)
                except Exception:
                    # fallback: executa um a um, engolindo erros pontuais
                    for sql, params in lote:
                        try:
                            conn.execute(sql, params)
                        except Exception:
                            pass
        pending.clear()

    SQL_ENV = """INSERT INTO pedidos
                 (numero,data_pedido,categoria,status,cliente,total,pagamento,forma_pagamento,
                  transportadora,ativo,enviado_em,status_ao_enviar)
                 VALUES (?,?,?,?,?,?,?,?,?,0,?,?)"""
    SQL_ATV = """INSERT INTO pedidos
                 (numero,data_pedido,categoria,status,cliente,total,pagamento,forma_pagamento,transportadora)
                 VALUES (?,?,?,?,?,?,?,?,?)"""

    def _coletar(o):
        """Decide o destino de um pedido da API e acumula os inserts necessários."""
        nonlocal salvos, duplicados, ja_enviados
        numero = str(o.get("number", "")).strip()
        if not numero:
            return

        # Pedido já existe no banco → não reprocessa (só completa itens se faltarem)
        if numero in existentes:
            duplicados += 1
            if numero not in com_itens:
                produtos = o.get("products") or []
                if produtos:
                    pending.extend(_build_item_stmts(numero, produtos))
                    com_itens.add(numero)
            return

        # Pedido novo → faz o parse completo (cancelados retornam None e são ignorados)
        c = _parse_order_fields(o, hoje)
        if not c:
            return
        existentes.add(numero)
        produtos = c["produtos"]

        if c["is_env"]:
            pending.append((SQL_ENV, (
                numero, c["data_str"], "Normal", c["status"], c["cliente"], c["total"],
                "Recebido", c["forma_pagamento"], c["transportadora"],
                hoje.isoformat(), c["status"])))
            ja_enviados += 1
        else:
            pending.append((SQL_ATV, (
                numero, c["data_str"], "Normal", c["status"], c["cliente"], c["total"],
                "Recebido", c["forma_pagamento"], c["transportadora"])))
            salvos += 1

        if produtos:
            pending.extend(_build_item_stmts(numero, produtos))
            com_itens.add(numero)

    # ── Fase 1: payment_status=paid (incremental via updated_at_min) ────────────
    page = 1
    while True:
        params = f"payment_status=paid&per_page=200&page={page}"
        if iso_q:
            params += f"&updated_at_min={iso_q}"
        url = f"https://api.nuvemshop.com.br/v1/{store_id}/orders?{params}"
        req = ureq.Request(url, headers=headers)
        try:
            with ureq.urlopen(req, timeout=20) as resp:
                orders = json.loads(resp.read())
        except Exception as e:
            return {"erro": str(e)}
        if not orders:
            break
        for o in orders:
            _coletar(o)
        page += 1
        if len(orders) < 200:
            break
    _marco("fase1_paid")

    # ── Fase 2: payment_status=authorized (incremental via updated_at_min) ──────
    page = 1
    while True:
        params = f"payment_status=authorized&per_page=200&page={page}"
        if iso_q:
            params += f"&updated_at_min={iso_q}"
        url = f"https://api.nuvemshop.com.br/v1/{store_id}/orders?{params}"
        req = ureq.Request(url, headers=headers)
        try:
            with ureq.urlopen(req, timeout=20) as resp:
                orders_auth = json.loads(resp.read())
        except Exception:
            break  # falha silenciosa — não bloqueia o sync principal
        if not orders_auth:
            break
        for o in orders_auth:
            _coletar(o)
        page += 1
        if len(orders_auth) < 200:
            break
    _marco("fase2_authorized")

    # Grava tudo que foi coletado nas fases 1 e 2 (poucos lotes)
    _flush()
    _marco("gravacao_banco")

    # ── Fase 3: cancelados na NuvemShop → marca como cancelado no banco ──────────
    reconciliados = 0
    try:
        agora_rec = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with get_conn() as conn:
            ativos_db = conn.execute(
                "SELECT id, numero FROM pedidos WHERE ativo=1"
            ).fetchall()
        ativos_map = {row["numero"]: row["id"] for row in ativos_db}

        if ativos_map:
            cancelados_nv = set()
            pg = 1
            while True:
                p_can = f"status=cancelled&per_page=200&page={pg}"
                if iso_q:
                    p_can += f"&updated_at_min={iso_q}"
                url_can = f"https://api.nuvemshop.com.br/v1/{store_id}/orders?{p_can}"
                req_can = ureq.Request(url_can, headers=headers)
                try:
                    with ureq.urlopen(req_can, timeout=20) as r:
                        orders_can = json.loads(r.read())
                except Exception:
                    break
                for o in orders_can:
                    cancelados_nv.add(str(o.get("number", "")))
                if len(orders_can) < 200:
                    break
                pg += 1

            updates = [
                ("UPDATE pedidos SET ativo=0, enviado_em=?, status_ao_enviar='Cancelado' WHERE id=?",
                 (agora_rec, pid))
                for numero, pid in ativos_map.items() if numero in cancelados_nv
            ]
            if updates:
                with get_conn() as conn:
                    for i in range(0, len(updates), 100):
                        try:
                            conn.execute_batch(updates[i:i + 100])
                        except Exception:
                            pass
                reconciliados = len(updates)
    except Exception:
        pass
    _marco("fase3_cancelados")

    # ── Desconto automático de estoque para pedidos novos ───────────────────
    try:
        _processar_estoque_pedidos()
    except Exception:
        pass
    _marco("estoque")

    agora = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S")
    with get_conn() as conn:
        conn.execute(
            "INSERT OR REPLACE INTO config (chave, valor) VALUES ('ultima_sincronizacao', ?)",
            (agora,),
        )

    return {"salvos": salvos, "duplicados": duplicados, "ja_enviados": ja_enviados,
            "reconciliados": reconciliados, "total": salvos + ja_enviados,
            "tempos": tempos}


@app.route("/api/sincronizar", methods=["POST"])
@login_required
def sincronizar_manual():
    try:
        resultado = sincronizar_nuvemshop()
    except Exception as e:
        resultado = {"erro": str(e)}
    return jsonify(resultado)


@app.route("/api/sync/status")
@login_required
def sync_status():
    store_id, _ = _get_nv_credentials()
    with get_conn() as conn:
        row = conn.execute("SELECT valor FROM config WHERE chave='ultima_sincronizacao'").fetchone()
    return jsonify({
        "configurado": bool(store_id),
        "ultima_sincronizacao": row["valor"] if row else None,
        "running": False,
        "ultimo_resultado": None,
    })


# ── NuvemShop OAuth callback ─────────────────────────────────────────────────

@app.route("/nuvemshop/callback")
def nuvemshop_callback():
    code     = request.args.get("code", "")
    store_id = request.args.get("store_id", "")

    if not code:
        return "<h2>Erro: code não recebido.</h2><pre>" + str(dict(request.args)) + "</pre>"

    import urllib.request, urllib.parse
    client_id     = os.getenv("NUVEMSHOP_CLIENT_ID", "")
    client_secret = os.getenv("NUVEMSHOP_CLIENT_SECRET", "")
    base = request.url_root.rstrip("/")
    if base.startswith("http://") and "ngrok" in base:
        base = base.replace("http://", "https://", 1)
    redirect_uri = base + "/nuvemshop/callback"

    payload = urllib.parse.urlencode({
        "client_id":     client_id,
        "client_secret": client_secret,
        "grant_type":    "authorization_code",
        "code":          code,
        "redirect_uri":  redirect_uri,
    }).encode()

    try:
        req = urllib.request.Request(
            "https://www.nuvemshop.com.br/apps/authorize/token",
            data=payload,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            token_data = json.loads(resp.read())

        if "error" in token_data:
            raise Exception(f"{token_data['error']}: {token_data.get('error_description','')}")

        access_token = token_data.get("access_token", "")
        tid          = str(token_data.get("user_id", store_id))

        # Salva credenciais no banco de dados (Vercel não tem sistema de arquivos persistente)
        with get_conn() as conn:
            conn.execute(
                "INSERT OR REPLACE INTO config (chave, valor) VALUES ('nuvemshop_store_id', ?)",
                (tid,),
            )
            conn.execute(
                "INSERT OR REPLACE INTO config (chave, valor) VALUES ('nuvemshop_access_token', ?)",
                (access_token,),
            )

        # Registra webhooks automaticamente após OAuth
        _registrar_webhooks_nuvemshop(tid, access_token, request.url_root.rstrip("/"))

        return f"""<html><body style="font-family:sans-serif;background:#0d1117;color:#e6edf3;padding:2rem">
            <h2 style="color:#3fb950">&#10003; Conectado com sucesso!</h2>
            <p><b>Store ID:</b> {tid}</p>
            <p><b>Token:</b> {access_token[:12]}…</p>
            <p>Salvo automaticamente. Pode fechar esta aba e voltar para a ferramenta.</p>
        </body></html>"""

    except Exception as e:
        return f"""<html><body style="font-family:sans-serif;background:#0d1117;color:#e6edf3;padding:2rem">
            <h2 style="color:#f85149">Erro ao trocar token</h2>
            <pre style="color:#f0883e">{e}</pre>
            <hr style="border-color:#30363d">
            <p style="color:#8b949e;font-size:.85rem"><b>Code recebido:</b> {code[:20]}…</p>
            <p style="color:#8b949e;font-size:.85rem"><b>redirect_uri usado:</b> {redirect_uri}</p>
        </body></html>"""


# ── NuvemShop Webhooks ────────────────────────────────────────────────────────

def _registrar_webhooks_nuvemshop(store_id, token, base_url):
    """Registra os webhooks de pedidos na NuvemShop (idempotente)."""
    import urllib.request as ureq
    headers = {
        "Authentication": f"bearer {token}",
        "User-Agent": "GS Mantos Interno (contato@gsmantos.com.br)",
        "Content-Type": "application/json",
    }

    # Busca webhooks já registrados para evitar duplicatas
    try:
        req = ureq.Request(
            f"https://api.nuvemshop.com.br/v1/{store_id}/webhooks",
            headers=headers,
        )
        with ureq.urlopen(req, timeout=15) as resp:
            existentes = json.loads(resp.read())
        urls_existentes = {wh.get("url", "") for wh in existentes}
    except Exception:
        urls_existentes = set()

    eventos = [
        ("orders/paid",      f"{base_url}/nuvemshop/webhooks/orders/paid"),
        ("orders/fulfilled", f"{base_url}/nuvemshop/webhooks/orders/fulfilled"),
        ("orders/cancelled", f"{base_url}/nuvemshop/webhooks/orders/cancelled"),
    ]

    for evento, url in eventos:
        if url in urls_existentes:
            continue
        try:
            payload = json.dumps({"event": evento, "url": url}).encode("utf-8")
            req = ureq.Request(
                f"https://api.nuvemshop.com.br/v1/{store_id}/webhooks",
                data=payload,
                headers=headers,
                method="POST",
            )
            ureq.urlopen(req, timeout=15)
        except Exception:
            pass


@app.route("/nuvemshop/webhooks/orders/paid", methods=["POST"])
def webhook_order_paid():
    """Recebe notificação da NuvemShop quando um pedido é pago."""
    try:
        order = request.get_json(force=True) or {}
    except Exception:
        return jsonify({"ok": False}), 400

    if not order:
        return jsonify({"ok": True}), 200

    with get_conn() as conn:
        resultado = _processar_order(conn, order)
        if resultado == "salvo":
            _sync_personalizacoes(conn)
            try:
                _processar_estoque_pedidos()
            except Exception:
                pass

    return jsonify({"ok": True, "resultado": resultado}), 200


@app.route("/nuvemshop/webhooks/orders/cancelled", methods=["POST"])
def webhook_order_cancelled():
    """Recebe notificação da NuvemShop quando um pedido é cancelado."""
    try:
        order = request.get_json(force=True) or {}
    except Exception:
        return jsonify({"ok": False}), 400

    numero = str(order.get("number", "")).strip()
    if not numero:
        return jsonify({"ok": True}), 200

    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        row = conn.execute(
            "SELECT id FROM pedidos WHERE numero = ? AND ativo = 1", (numero,)
        ).fetchone()
        if row:
            conn.execute(
                "UPDATE pedidos SET ativo=0, enviado_em=?, status_ao_enviar='Cancelado' WHERE id=?",
                (agora, row["id"]),
            )

    return jsonify({"ok": True}), 200


@app.route("/nuvemshop/webhooks/orders/fulfilled", methods=["POST"])
def webhook_order_fulfilled():
    """Recebe notificação da NuvemShop quando um pedido é enviado."""
    try:
        order = request.get_json(force=True) or {}
    except Exception:
        return jsonify({"ok": False}), 400

    numero = str(order.get("number", "")).strip()
    if not numero:
        return jsonify({"ok": True}), 200

    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        row = conn.execute(
            "SELECT id, status FROM pedidos WHERE numero = ? AND ativo = 1", (numero,)
        ).fetchone()
        if row:
            conn.execute(
                "UPDATE pedidos SET ativo=0, enviado_em=?, status_ao_enviar=? WHERE id=?",
                (agora, row["status"], row["id"]),
            )

    return jsonify({"ok": True}), 200


@app.route("/api/nuvemshop/registrar-webhooks", methods=["POST"])
@login_required
def api_registrar_webhooks():
    """Registra manualmente os webhooks (caso o OAuth tenha sido feito antes desta feature)."""
    store_id, token = _get_nv_credentials()
    if not store_id or not token:
        return jsonify({"erro": "NuvemShop não configurada"}), 400
    base_url = request.url_root.rstrip("/")
    try:
        _registrar_webhooks_nuvemshop(store_id, token, base_url)
        return jsonify({"ok": True, "mensagem": "Webhooks registrados com sucesso"})
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route("/api/nuvemshop/limpar-cancelados", methods=["POST"])
@login_required
def limpar_cancelados():
    """Remove do banco todos os pedidos ativos que estão cancelados na NuvemShop."""
    import urllib.request as ureq
    store_id, token = _get_nv_credentials()
    if not store_id or not token:
        return jsonify({"erro": "NuvemShop não configurada"}), 400

    headers = _nuvemshop_headers()
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Busca todos os pedidos cancelados na NuvemShop
    cancelados_nv = set()
    pg = 1
    while True:
        url = f"https://api.nuvemshop.com.br/v1/{store_id}/orders?status=cancelled&per_page=200&page={pg}"
        req = ureq.Request(url, headers=headers)
        try:
            with ureq.urlopen(req, timeout=20) as r:
                orders = json.loads(r.read())
        except Exception as e:
            return jsonify({"erro": str(e)}), 500
        for o in orders:
            cancelados_nv.add(str(o.get("number", "")))
        if len(orders) < 200:
            break
        pg += 1

    if not cancelados_nv:
        return jsonify({"removidos": 0, "mensagem": "Nenhum pedido cancelado encontrado na NuvemShop"})

    # Remove do banco os que estão ativos mas cancelados na NuvemShop (em lote)
    with get_conn() as conn:
        ativos = conn.execute("SELECT id, numero FROM pedidos WHERE ativo=1").fetchall()
    updates = [
        ("UPDATE pedidos SET ativo=0, enviado_em=?, status_ao_enviar='Cancelado' WHERE id=?",
         (agora, row["id"]))
        for row in ativos if row["numero"] in cancelados_nv
    ]
    removidos = len(updates)
    if updates:
        with get_conn() as conn:
            for i in range(0, len(updates), 100):
                conn.execute_batch(updates[i:i + 100])

    return jsonify({"removidos": removidos, "mensagem": f"{removidos} pedido(s) cancelado(s) removido(s) da fila"})


# ── Helpers — Estoque ────────────────────────────────────────────────────────

def _build_sync_query(produto_nome, variante_label, sku_origem, skip_vid):
    """Monta a query de busca de registros para sincronização.

    Regra de segurança para SKUs diferentes:
    - Se o registro ORIGEM tem SKU preenchido: só sincroniza com quem tem o MESMO SKU.
      Isso evita que produtos com mesmo nome mas SKUs distintos (produtos diferentes)
      sejam sincronizados por engano.
    - Se o registro ORIGEM não tem SKU: sincroniza apenas com quem também não tem SKU
      (ou seja, nunca cruza SKU preenchido com não-preenchido).
    """
    q = """SELECT id, nv_variant_id, sku, quantity FROM sku_stock
           WHERE UPPER(TRIM(produto_nome)) = UPPER(TRIM(?))"""
    params = [produto_nome]

    if variante_label:
        q += " AND UPPER(TRIM(variante_label)) = UPPER(TRIM(?))"
        params.append(variante_label)

    if sku_origem:
        # Origem tem SKU → só bate com quem tem o mesmo SKU
        q += " AND UPPER(TRIM(COALESCE(sku,''))) = UPPER(TRIM(?))"
        params.append(sku_origem)
    else:
        # Origem sem SKU → só bate com quem também não tem SKU
        q += " AND (sku IS NULL OR TRIM(sku) = '')"

    if skip_vid:
        q += " AND nv_variant_id != ?"
        params.append(skip_vid)

    return q, params


def _sync_estoque_por_nome(conn, produto_nome, variante_label, nova_qty, agora,
                           skip_vid=None, sku_origem=None):
    """Sincroniza TODOS os registros elegíveis (mesmo nome + mesma SKU se houver) para nova_qty.
    Retorna quantos registros foram sincronizados além do original."""
    if not produto_nome:
        return 0
    q, params = _build_sync_query(produto_nome, variante_label, sku_origem, skip_vid)
    rows = conn.execute(q, params).fetchall()
    for r in rows:
        conn.execute(
            "UPDATE sku_stock SET quantity=?, updated_at=? WHERE id=?",
            (nova_qty, agora, r["id"])
        )
        conn.execute(
            """INSERT INTO sku_stock_movements
               (nv_variant_id, sku, tipo, quantidade, observacao, created_at)
               VALUES (?, ?, 'ajuste', ?, 'Sync automático por nome', ?)""",
            (r["nv_variant_id"], r["sku"], nova_qty, agora)
        )
    return len(rows)


def _deducao_sync_por_nome(conn, produto_nome, variante_label, qty_deduzida, agora,
                           skip_vid=None, pedido_nr=None, sku_origem=None):
    """Deduz qty_deduzida de TODOS os registros elegíveis (mesmo nome + mesma SKU se houver)."""
    if not produto_nome:
        return 0
    q, params = _build_sync_query(produto_nome, variante_label, sku_origem, skip_vid)
    rows = conn.execute(q, params).fetchall()
    for r in rows:
        nova = r["quantity"] - qty_deduzida
        conn.execute(
            "UPDATE sku_stock SET quantity=?, updated_at=? WHERE id=?",
            (nova, agora, r["id"])
        )
        conn.execute(
            """INSERT INTO sku_stock_movements
               (nv_variant_id, sku, tipo, quantidade, pedido_numero, observacao, created_at)
               VALUES (?, ?, 'saida_venda', ?, ?, 'Sync automático por nome', ?)""",
            (r["nv_variant_id"], r["sku"], qty_deduzida, pedido_nr, agora)
        )
    return len(rows)


# ── Routes — Estoque ─────────────────────────────────────────────────────────

@app.route("/api/estoque", methods=["GET"])
@app.route("/api/estoque/listar", methods=["GET"])
@login_required
def api_estoque_get():
    """Retorna todos os registros de estoque."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT s.*,
                   (SELECT COUNT(*) FROM sku_stock_movements m WHERE m.nv_variant_id = s.nv_variant_id) AS total_movimentos
            FROM sku_stock s
            ORDER BY s.updated_at DESC
        """).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/estoque/ajustar", methods=["POST"])
@login_required
def api_estoque_ajustar():
    """Ajuste manual de estoque: entrada, saida_manual ou ajuste (valor absoluto)."""
    data          = request.get_json() or {}
    nv_variant_id = data.get("nv_variant_id")
    nv_product_id = data.get("nv_product_id")
    sku           = (data.get("sku") or "").strip() or None
    tipo          = data.get("tipo", "entrada")        # entrada | saida_manual | ajuste
    quantidade    = data.get("quantidade", 0)
    min_quantity  = data.get("min_quantity")
    observacao    = data.get("observacao")
    produto_nome  = (data.get("produto_nome") or "").strip() or None
    variante_label= (data.get("variante_label") or "").strip() or None

    if not nv_variant_id and not sku:
        return jsonify({"erro": "nv_variant_id ou sku obrigatório"}), 400
    if tipo not in ("entrada", "saida_manual", "ajuste"):
        return jsonify({"erro": "tipo inválido"}), 400

    try:
        quantidade = int(quantidade)
    except (TypeError, ValueError):
        return jsonify({"erro": "quantidade inválida"}), 400

    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with get_conn() as conn:
        if nv_variant_id:
            existing = conn.execute(
                "SELECT id, quantity FROM sku_stock WHERE nv_variant_id=?", (nv_variant_id,)
            ).fetchone()
        else:
            existing = conn.execute(
                "SELECT id, quantity FROM sku_stock WHERE sku=? AND nv_variant_id IS NULL", (sku,)
            ).fetchone()

        if existing:
            if tipo == "ajuste":
                nova_qty = quantidade          # valor absoluto
            elif tipo == "entrada":
                nova_qty = existing["quantity"] + quantidade
            else:  # saida_manual
                nova_qty = existing["quantity"] - quantidade

            upd = "UPDATE sku_stock SET quantity=?, updated_at=?"
            params = [nova_qty, agora]
            if min_quantity is not None:
                upd += ", min_quantity=?"
                params.append(int(min_quantity))
            # Grava nomes se fornecidos
            if produto_nome:
                upd += ", produto_nome=?"
                params.append(produto_nome)
            if variante_label:
                upd += ", variante_label=?"
                params.append(variante_label)
            upd += " WHERE id=?"
            params.append(existing["id"])
            conn.execute(upd, params)
        else:
            # Cria novo registro
            if tipo == "ajuste":
                nova_qty = quantidade
            elif tipo == "entrada":
                nova_qty = quantidade
            else:
                nova_qty = -quantidade

            min_q = int(min_quantity) if min_quantity is not None else 3
            conn.execute(
                """INSERT INTO sku_stock
                   (nv_variant_id, nv_product_id, sku, quantity, min_quantity,
                    produto_nome, variante_label, updated_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                (nv_variant_id, nv_product_id, sku, nova_qty, min_q,
                 produto_nome, variante_label, agora),
            )

        # Registra movimento (exceto ajuste que só muda min)
        if tipo != "ajuste" or quantidade != 0:
            conn.execute(
                """INSERT INTO sku_stock_movements
                   (nv_variant_id, sku, tipo, quantidade, observacao, created_at)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (nv_variant_id, sku, tipo, abs(quantidade), observacao, agora),
            )

        # ── Sincroniza produtos com mesmo nome + mesma SKU ─────────────────
        synced = _sync_estoque_por_nome(
            conn, produto_nome, variante_label, nova_qty, agora,
            skip_vid=nv_variant_id, sku_origem=sku
        )

    return jsonify({"ok": True, "nova_quantidade": nova_qty, "sincronizados": synced})


@app.route("/api/estoque/sincronizar-catalogo", methods=["POST"])
@login_required
def api_estoque_sincronizar_catalogo():
    """Busca todos os produtos do NuvemShop, popula produto_nome+variante_label no sku_stock
    e sincroniza quantidades de registros com mesmo nome."""
    import urllib.request as ureq
    store_id, token = _get_nv_credentials()
    if not store_id or not token:
        return jsonify({"erro": "NuvemShop não configurada"}), 400

    headers  = _nuvemshop_headers()
    agora    = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    atualizados = 0
    sincronizados = 0

    # Busca todos os produtos
    all_products = []
    pg = 1
    while True:
        url = f"https://api.nuvemshop.com.br/v1/{store_id}/products?per_page=200&page={pg}"
        req = ureq.Request(url, headers=headers)
        try:
            with ureq.urlopen(req, timeout=30) as resp:
                batch = json.loads(resp.read())
        except Exception as e:
            return jsonify({"erro": str(e)}), 502
        if not batch:
            break
        all_products.extend(batch)
        pg += 1
        if len(batch) < 200:
            break

    with get_conn() as conn:
        for p in all_products:
            nome = _extract_text(p.get("name", "")) or ""
            for v in (p.get("variants") or []):
                vid = v.get("id")
                if not vid:
                    continue
                vals  = v.get("values") or []
                label = ", ".join(_extract_text(x) for x in vals if x)

                existing = conn.execute(
                    "SELECT id FROM sku_stock WHERE nv_variant_id=?", (vid,)
                ).fetchone()
                if existing:
                    conn.execute(
                        "UPDATE sku_stock SET produto_nome=?, variante_label=? WHERE id=?",
                        (nome, label, existing["id"])
                    )
                    atualizados += 1

        conn.commit()

        # Agora sincroniza quantidades: agrupa por (produto_nome, variante_label, SKU normalizado)
        # SKU nulo/vazio é tratado como grupo próprio: '' agrupa com ''
        # SKU preenchido só agrupa com o mesmo SKU — evita sync de produtos distintos
        grupos = conn.execute("""
            SELECT produto_nome, variante_label,
                   UPPER(TRIM(COALESCE(sku,''))) as sku_norm,
                   MAX(quantity) as max_qty
            FROM sku_stock
            WHERE produto_nome IS NOT NULL AND variante_label IS NOT NULL
            GROUP BY UPPER(TRIM(produto_nome)),
                     UPPER(TRIM(variante_label)),
                     UPPER(TRIM(COALESCE(sku,'')))
            HAVING COUNT(*) > 1
        """).fetchall()

        for g in grupos:
            sku_norm = g["sku_norm"]
            if sku_norm:
                rows = conn.execute("""
                    SELECT id, nv_variant_id, sku FROM sku_stock
                    WHERE UPPER(TRIM(produto_nome))  = UPPER(TRIM(?))
                      AND UPPER(TRIM(variante_label))= UPPER(TRIM(?))
                      AND UPPER(TRIM(COALESCE(sku,''))) = UPPER(TRIM(?))
                """, (g["produto_nome"], g["variante_label"], sku_norm)).fetchall()
            else:
                rows = conn.execute("""
                    SELECT id, nv_variant_id, sku FROM sku_stock
                    WHERE UPPER(TRIM(produto_nome))  = UPPER(TRIM(?))
                      AND UPPER(TRIM(variante_label))= UPPER(TRIM(?))
                      AND (sku IS NULL OR TRIM(sku) = '')
                """, (g["produto_nome"], g["variante_label"])).fetchall()

            for r in rows:
                conn.execute(
                    "UPDATE sku_stock SET quantity=?, updated_at=? WHERE id=?",
                    (g["max_qty"], agora, r["id"])
                )
                conn.execute("""INSERT INTO sku_stock_movements
                    (nv_variant_id, sku, tipo, quantidade, observacao, created_at)
                    VALUES (?, ?, 'ajuste', ?, 'Sync por nome — catálogo', ?)""",
                    (r["nv_variant_id"], r["sku"], g["max_qty"], agora))
                sincronizados += 1
        conn.commit()

    return jsonify({
        "ok": True,
        "produtos_catalogados": atualizados,
        "grupos_sincronizados": len(grupos),
        "registros_sincronizados": sincronizados,
    })


@app.route("/api/estoque/movimentos", methods=["GET"])
@login_required
def api_estoque_movimentos():
    """Histórico de movimentos. Filtros: nv_variant_id, limit."""
    vid   = request.args.get("nv_variant_id")
    limit = int(request.args.get("limit", 50))

    with get_conn() as conn:
        if vid:
            rows = conn.execute(
                """SELECT * FROM sku_stock_movements WHERE nv_variant_id=?
                   ORDER BY created_at DESC LIMIT ?""",
                (int(vid), limit),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM sku_stock_movements ORDER BY created_at DESC LIMIT ?",
                (limit,),
            ).fetchall()

    return jsonify([dict(r) for r in rows])


# ── Helper: extrai texto multilíngue da NuvemShop ────────────────────────────

def _extract_text(obj):
    if isinstance(obj, str):
        return obj
    if isinstance(obj, dict):
        return (obj.get("pt") or obj.get("pt-BR") or obj.get("es")
                or obj.get("en") or next(iter(obj.values()), "") or "")
    return str(obj) if obj is not None else ""


# ── Routes — Custos SKU ───────────────────────────────────────────────────────

@app.route("/custos-sku")
@login_required
def custos_sku_page():
    return render_template("custos_sku.html")


@app.route("/api/sku-costs", methods=["GET"])
@login_required
def api_sku_costs_get():
    sku             = request.args.get("sku")
    variant_ids_str = request.args.get("nv_variant_ids")

    with get_conn() as conn:
        if sku:
            rows = conn.execute(
                "SELECT * FROM sku_costs WHERE sku=? ORDER BY effective_from DESC",
                (sku,)
            ).fetchall()

        elif variant_ids_str:
            try:
                ids = [int(x.strip()) for x in variant_ids_str.split(",") if x.strip()]
            except ValueError:
                return jsonify([])
            if not ids:
                return jsonify([])
            ph = ",".join("?" * len(ids))
            rows = conn.execute(f"""
                SELECT s.* FROM sku_costs s
                WHERE s.nv_variant_id IN ({ph})
                  AND (s.effective_to IS NULL OR s.effective_to >= date('now'))
                GROUP BY s.nv_variant_id
                HAVING MAX(s.effective_from)
            """, ids).fetchall()

        else:
            # Todos os custos vigentes (um por variante / sku)
            # CRÍTICO: GROUP BY COALESCE para não colapsar variantes com mesmo sku
            rows = conn.execute("""
                SELECT s.*,
                       (SELECT COUNT(*) FROM sku_costs s2 WHERE s2.sku=s.sku) AS history_count
                FROM sku_costs s
                WHERE s.effective_to IS NULL OR s.effective_to >= date('now')
                GROUP BY COALESCE(CAST(s.nv_variant_id AS TEXT), s.sku)
                HAVING MAX(s.effective_from)
                ORDER BY s.sku
            """).fetchall()

    return jsonify([dict(r) for r in rows])


@app.route("/api/sku-costs", methods=["POST"])
@login_required
def api_sku_costs_post():
    data = request.get_json() or {}
    sku            = (data.get("sku") or "").strip()
    name           = data.get("name")
    type_          = data.get("type", "product")
    cost           = data.get("cost", 0)
    effective_from = data.get("effective_from") or date.today().isoformat()
    effective_to   = data.get("effective_to")
    notes          = data.get("notes")
    nv_variant_id  = data.get("nv_variant_id")
    nv_product_id  = data.get("nv_product_id")

    if not sku:
        return jsonify({"erro": "SKU obrigatório"}), 400
    try:
        cost = float(cost)
        if cost < 0:
            raise ValueError
    except (TypeError, ValueError):
        return jsonify({"erro": "Custo inválido"}), 400
    if type_ not in ("product", "brinde"):
        type_ = "product"

    with get_conn() as conn:
        # Fecha registro anterior (preserva histórico)
        if nv_variant_id:
            conn.execute(
                """UPDATE sku_costs SET effective_to = date(?, '-1 day')
                   WHERE nv_variant_id=? AND effective_to IS NULL AND effective_from < ?""",
                (effective_from, nv_variant_id, effective_from),
            )
        else:
            conn.execute(
                """UPDATE sku_costs SET effective_to = date(?, '-1 day')
                   WHERE sku=? AND nv_variant_id IS NULL
                     AND effective_to IS NULL AND effective_from < ?""",
                (effective_from, sku, effective_from),
            )
        # Insere novo
        conn.execute(
            """INSERT INTO sku_costs
               (sku, name, type, cost, effective_from, effective_to, notes, nv_variant_id, nv_product_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (sku, name, type_, cost, effective_from, effective_to, notes, nv_variant_id, nv_product_id),
        )

    return jsonify({"ok": True})


@app.route("/api/estoque/entrada-lote", methods=["POST"])
@login_required
def api_estoque_entrada_lote():
    """Entrada de estoque com preço de compra — calcula Custo Médio Ponderado (CMP).

    Body JSON:
        nv_variant_id  int
        nv_product_id  int  (opcional)
        sku            str
        quantidade     int   — quantidade do novo lote
        preco_compra   float — preço pago por unidade neste lote
        produto_nome   str  (opcional — para sync por nome)
        variante_label str  (opcional)
    """
    data          = request.get_json() or {}
    nv_variant_id = data.get("nv_variant_id")
    nv_product_id = data.get("nv_product_id")
    sku           = (data.get("sku") or "").strip() or None
    quantidade    = data.get("quantidade", 0)
    preco_compra  = data.get("preco_compra", 0)
    produto_nome  = (data.get("produto_nome") or "").strip() or None
    variante_label= (data.get("variante_label") or "").strip() or None

    if not nv_variant_id and not sku:
        return jsonify({"erro": "nv_variant_id ou sku obrigatório"}), 400

    # sku_costs.sku é NOT NULL — usa nv_variant_id como fallback se não tiver SKU
    sku_custo = sku or (f"VID-{nv_variant_id}" if nv_variant_id else None)
    if not sku_custo:
        return jsonify({"erro": "SKU não encontrado para esta variante"}), 400
    try:
        quantidade   = int(quantidade)
        preco_compra = float(preco_compra)
        if quantidade <= 0:
            raise ValueError("quantidade deve ser positiva")
        if preco_compra < 0:
            raise ValueError("preço não pode ser negativo")
    except (TypeError, ValueError) as e:
        return jsonify({"erro": str(e)}), 400

    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    hoje  = date.today().isoformat()

    with get_conn() as conn:
        # ── Estoque atual ────────────────────────────────────────────
        if nv_variant_id:
            stock = conn.execute(
                "SELECT id, quantity, produto_nome, variante_label FROM sku_stock WHERE nv_variant_id=?",
                (nv_variant_id,)
            ).fetchone()
        else:
            stock = conn.execute(
                "SELECT id, quantity, produto_nome, variante_label FROM sku_stock WHERE sku=? AND nv_variant_id IS NULL",
                (sku,)
            ).fetchone()

        qty_atual = int(stock["quantity"]) if stock and stock["quantity"] else 0

        # ── Custo atual (vigente) ────────────────────────────────────
        if nv_variant_id:
            custo_row = conn.execute(
                """SELECT cost FROM sku_costs
                   WHERE nv_variant_id=? AND effective_to IS NULL
                   ORDER BY effective_from DESC LIMIT 1""",
                (nv_variant_id,)
            ).fetchone()
        else:
            custo_row = conn.execute(
                """SELECT cost FROM sku_costs
                   WHERE sku=? AND nv_variant_id IS NULL AND effective_to IS NULL
                   ORDER BY effective_from DESC LIMIT 1""",
                (sku,)
            ).fetchone()

        custo_atual = float(custo_row["cost"]) if custo_row else 0.0

        # ── Calcula CMP ──────────────────────────────────────────────
        # Se não tem estoque ou custo, o CMP é o preço do lote atual
        if qty_atual <= 0 or custo_atual == 0:
            novo_cmp = preco_compra
        else:
            total_valor = (qty_atual * custo_atual) + (quantidade * preco_compra)
            nova_qty_total = qty_atual + quantidade
            novo_cmp = round(total_valor / nova_qty_total, 4)

        nova_qty = qty_atual + quantidade

        # ── Atualiza estoque ─────────────────────────────────────────
        if stock:
            upd = "UPDATE sku_stock SET quantity=?, updated_at=?"
            params = [nova_qty, agora]
            if produto_nome:
                upd += ", produto_nome=?"; params.append(produto_nome)
            if variante_label:
                upd += ", variante_label=?"; params.append(variante_label)
            upd += " WHERE id=?"; params.append(stock["id"])
            conn.execute(upd, params)
        else:
            min_q = 3
            conn.execute(
                """INSERT INTO sku_stock
                   (nv_variant_id, nv_product_id, sku, quantity, min_quantity,
                    produto_nome, variante_label, updated_at)
                   VALUES (?,?,?,?,?,?,?,?)""",
                (nv_variant_id, nv_product_id, sku, nova_qty, min_q,
                 produto_nome, variante_label, agora)
            )

        # ── Registra movimento com preço de compra ───────────────────
        conn.execute(
            """INSERT INTO sku_stock_movements
               (nv_variant_id, sku, tipo, quantidade, preco_compra, observacao, created_at)
               VALUES (?,?,?,?,?,?,?)""",
            (nv_variant_id, sku, "entrada", quantidade, preco_compra,
             f"Lote: {quantidade}un × R${preco_compra:.2f} | CMP: R${novo_cmp:.2f}", agora)
        )

        # ── Atualiza custo com o novo CMP ────────────────────────────
        # Fecha registro anterior
        if nv_variant_id:
            conn.execute(
                """UPDATE sku_costs SET effective_to=date(?, '-1 day')
                   WHERE nv_variant_id=? AND effective_to IS NULL AND effective_from < ?""",
                (hoje, nv_variant_id, hoje)
            )
        conn.execute(
            """UPDATE sku_costs SET effective_to=date(?, '-1 day')
               WHERE sku=? AND effective_to IS NULL AND effective_from < ?""",
            (hoje, sku_custo, hoje)
        )
        # Insere novo registro com CMP calculado
        nome_custo = produto_nome or (stock["produto_nome"] if stock else None)
        conn.execute(
            """INSERT INTO sku_costs
               (sku, name, type, cost, effective_from, notes, nv_variant_id, nv_product_id)
               VALUES (?,?,?,?,?,?,?,?)""",
            (sku_custo, nome_custo, "product", novo_cmp, hoje,
             f"CMP: {qty_atual}un×R${custo_atual:.2f} + {quantidade}un×R${preco_compra:.2f} = {nova_qty}un×R${novo_cmp:.2f}",
             nv_variant_id, nv_product_id)
        )

        # ── Sync por nome + SKU ──────────────────────────────────────
        _sync_estoque_por_nome(
            conn, produto_nome or (stock["produto_nome"] if stock else None),
            variante_label or (stock["variante_label"] if stock else None),
            nova_qty, agora, skip_vid=nv_variant_id, sku_origem=sku
        )

        conn.commit()

    return jsonify({
        "ok": True,
        "qty_anterior": qty_atual,
        "custo_anterior": custo_atual,
        "nova_quantidade": nova_qty,
        "novo_cmp": novo_cmp,
        "preco_lote": preco_compra,
        "quantidade_lote": quantidade,
    })


@app.route("/api/nuvemshop/produtos", methods=["GET"])
@login_required
def api_nuvemshop_produtos():
    import urllib.request as ureq
    store_id, token = _get_nv_credentials()
    if not store_id or not token:
        return jsonify({"erro": "NuvemShop não configurada"}), 400

    headers     = _nuvemshop_headers()
    sort_by     = request.args.get("sort_by", "newest")
    category_id = request.args.get("category_id", "")

    sort_map = {
        "newest":     "created-descending",
        "oldest":     "created-ascending",
        "most_sold":  "sales-descending",
        "least_sold": "sales-ascending",
    }
    nv_sort = sort_map.get(sort_by, "created-descending")

    products = []
    pg = 1
    while True:
        params = f"sort_by={nv_sort}&per_page=200&page={pg}"
        if category_id:
            params += f"&category_id={urllib.parse.quote(str(category_id))}"
        url = f"https://api.nuvemshop.com.br/v1/{store_id}/products?{params}"
        req = ureq.Request(url, headers=headers)
        try:
            with ureq.urlopen(req, timeout=30) as resp:
                batch = json.loads(resp.read())
        except Exception as e:
            return jsonify({"erro": str(e)}), 502

        for p in batch:
            name = _extract_text(p.get("name", ""))

            # Imagem principal
            imgs      = p.get("images") or []
            image_url = None
            if imgs:
                img0      = imgs[0]
                image_url = img0.get("src") if isinstance(img0, dict) else None

            # Categorias → lista de IDs
            raw_cats = p.get("categories") or []
            cat_ids  = []
            for c in raw_cats:
                if isinstance(c, dict):
                    cat_ids.append(c.get("id"))
                elif isinstance(c, int):
                    cat_ids.append(c)

            # Variantes
            variants = []
            for v in (p.get("variants") or []):
                v_vals   = v.get("values") or []
                v_labels = [_extract_text(x) for x in v_vals]
                v_img    = v.get("image") or {}
                v_img_url = (v_img.get("src") if isinstance(v_img, dict) else None) or image_url
                try:
                    price = float(str(v.get("price") or 0).replace(",", "."))
                except Exception:
                    price = 0.0
                variants.append({
                    "id":       v.get("id"),
                    "sku":      v.get("sku"),
                    "values":   v_labels,
                    "price":    price,
                    "imageUrl": v_img_url,
                })

            products.append({
                "id":           p.get("id"),
                "name":         name,
                "imageUrl":     image_url,
                "created_at":   p.get("created_at", ""),
                "category_ids": cat_ids,
                "sales_count":  0,
                "variants":     variants,
            })

        pg += 1
        if len(batch) < 200:
            break

    return jsonify(products)


@app.route("/api/nuvemshop/categorias", methods=["GET"])
@login_required
def api_nuvemshop_categorias():
    import urllib.request as ureq
    store_id, token = _get_nv_credentials()
    if not store_id or not token:
        return jsonify([])

    headers = _nuvemshop_headers()
    url = f"https://api.nuvemshop.com.br/v1/{store_id}/categories?per_page=200"
    req = ureq.Request(url, headers=headers)
    try:
        with ureq.urlopen(req, timeout=15) as resp:
            cats = json.loads(resp.read())
    except Exception:
        return jsonify([])

    result = []
    for c in cats:
        parent    = c.get("parent") or {}
        parent_id = parent.get("id") if isinstance(parent, dict) else None
        result.append({
            "id":        c.get("id"),
            "name":      _extract_text(c.get("name", "")),
            "parent_id": parent_id,
        })
    return jsonify(result)


# ── Routes — Custo de Personalização ─────────────────────────────────────────

PERS_COST_KEYS = (
    "pers_custo_nome",      # R$ por nome impresso
    "pers_custo_numero",    # R$ por número impresso
    "pers_custo_escudo",    # R$ por escudo/logo (opcional)
    "pers_ativo",           # "1" se o módulo está ativo
)

@app.route("/api/personalizacao/config", methods=["GET"])
@login_required
def get_pers_config():
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT chave, valor FROM config WHERE chave IN (?,?,?,?)",
            PERS_COST_KEYS,
        ).fetchall()
    cfg = {r["chave"]: r["valor"] for r in rows}
    return jsonify({
        "ativo":         cfg.get("pers_ativo", "1") == "1",
        "custo_nome":    float(cfg.get("pers_custo_nome",   0) or 0),
        "custo_numero":  float(cfg.get("pers_custo_numero", 0) or 0),
        "custo_escudo":  float(cfg.get("pers_custo_escudo", 0) or 0),
    })


@app.route("/api/personalizacao/config", methods=["POST"])
@login_required
def set_pers_config():
    data          = request.get_json() or {}
    custo_nome    = max(0.0, float(data.get("custo_nome",   0) or 0))
    custo_numero  = max(0.0, float(data.get("custo_numero", 0) or 0))
    custo_escudo  = max(0.0, float(data.get("custo_escudo", 0) or 0))
    ativo         = "1" if data.get("ativo", True) else "0"
    with get_conn() as conn:
        for chave, valor in [
            ("pers_custo_nome",   str(custo_nome)),
            ("pers_custo_numero", str(custo_numero)),
            ("pers_custo_escudo", str(custo_escudo)),
            ("pers_ativo",        ativo),
        ]:
            conn.execute(
                "INSERT OR REPLACE INTO config (chave, valor) VALUES (?, ?)",
                (chave, valor),
            )
    return jsonify({"ok": True})


@app.route("/api/personalizacao/calcular", methods=["GET"])
@login_required
def calcular_custo_pers():
    """Calcula custo adicional de personalização para uma combinação nomes/números/escudos."""
    qtd_nomes   = max(0, int(request.args.get("nomes",   0) or 0))
    qtd_numeros = max(0, int(request.args.get("numeros", 0) or 0))
    qtd_escudos = max(0, int(request.args.get("escudos", 0) or 0))
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT chave, valor FROM config WHERE chave IN (?,?,?,?)",
            PERS_COST_KEYS,
        ).fetchall()
    cfg = {r["chave"]: float(r["valor"] or 0) for r in rows}
    custo_nome   = cfg.get("pers_custo_nome",   0)
    custo_numero = cfg.get("pers_custo_numero", 0)
    custo_escudo = cfg.get("pers_custo_escudo", 0)
    total = qtd_nomes * custo_nome + qtd_numeros * custo_numero + qtd_escudos * custo_escudo
    return jsonify({
        "qtd_nomes":    qtd_nomes,
        "qtd_numeros":  qtd_numeros,
        "qtd_escudos":  qtd_escudos,
        "custo_nome":   custo_nome,
        "custo_numero": custo_numero,
        "custo_escudo": custo_escudo,
        "total":        round(total, 2),
        "breakdown": [
            {"item": "Nome(s)",   "qtd": qtd_nomes,   "unit": custo_nome,   "subtotal": round(qtd_nomes   * custo_nome,   2)},
            {"item": "Número(s)", "qtd": qtd_numeros, "unit": custo_numero, "subtotal": round(qtd_numeros * custo_numero, 2)},
            {"item": "Escudo(s)", "qtd": qtd_escudos, "unit": custo_escudo, "subtotal": round(qtd_escudos * custo_escudo, 2)},
        ],
    })


# ── Routes — Custo de Personalização por Variante ────────────────────────────

@app.route("/api/sku-pers-pricing", methods=["GET"])
@login_required
def api_get_sku_pers_pricing():
    """Retorna configuração de personalização para uma ou mais variantes (nv_variant_ids separados por vírgula)."""
    nv_variant_ids_str = request.args.get("nv_variant_ids", "")
    if not nv_variant_ids_str:
        return jsonify([])
    try:
        ids = [int(x.strip()) for x in nv_variant_ids_str.split(",") if x.strip()]
    except ValueError:
        return jsonify([])
    ph = ",".join("?" * len(ids))
    with get_conn() as conn:
        rows = conn.execute(
            f"SELECT * FROM sku_pers_pricing WHERE nv_variant_id IN ({ph})", ids
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/sku-pers-pricing", methods=["POST"])
@login_required
def api_post_sku_pers_pricing():
    """Salva (upsert) configuração de personalização para uma variante."""
    data = request.get_json() or {}
    nv_variant_id = data.get("nv_variant_id")
    sku           = data.get("sku", "")
    custo_nome    = max(0.0, float(data.get("custo_nome",   0) or 0))
    custo_numero  = max(0.0, float(data.get("custo_numero", 0) or 0))
    custo_escudo  = max(0.0, float(data.get("custo_escudo", 0) or 0))
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        conn.execute(
            """INSERT INTO sku_pers_pricing (nv_variant_id, sku, custo_nome, custo_numero, custo_escudo, updated_at)
               VALUES (?, ?, ?, ?, ?, ?)
               ON CONFLICT(nv_variant_id) DO UPDATE SET
                   sku=excluded.sku, custo_nome=excluded.custo_nome,
                   custo_numero=excluded.custo_numero, custo_escudo=excluded.custo_escudo,
                   updated_at=excluded.updated_at""",
            (nv_variant_id, sku, custo_nome, custo_numero, custo_escudo, agora)
        )
    return jsonify({"ok": True})


# ── Routes — Projeção de Faturamento por Estoque ─────────────────────────────

@app.route("/projecao-estoque")
@login_required
def projecao_estoque_page():
    return render_template("projecao_estoque.html")


@app.route("/api/projecao-estoque")
@login_required
def api_projecao_estoque():
    """Retorna estoque + custo vigente por variante para projeção de faturamento."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT
                s.id,
                s.nv_variant_id,
                s.nv_product_id,
                s.sku,
                s.quantity,
                s.min_quantity,
                s.updated_at,
                sc.cost,
                sc.name   AS cost_name,
                sc.type   AS cost_type
            FROM sku_stock s
            LEFT JOIN sku_costs sc ON sc.id = (
                SELECT id FROM sku_costs sc2
                WHERE (
                    (s.nv_variant_id IS NOT NULL AND sc2.nv_variant_id = s.nv_variant_id)
                    OR (s.nv_variant_id IS NULL  AND sc2.sku = s.sku AND sc2.nv_variant_id IS NULL)
                )
                AND (sc2.effective_to IS NULL OR sc2.effective_to >= date('now'))
                ORDER BY sc2.effective_from DESC
                LIMIT 1
            )
            ORDER BY s.quantity DESC
        """).fetchall()
    return jsonify([dict(r) for r in rows])


# ── Routes — Dashboard de Compras ────────────────────────────────────────────

@app.route("/compras")
@login_required
def compras_page():
    return render_template("compras.html")


@app.route("/api/compras/dia")
@login_required
def api_compras_dia():
    """Retorna produtos com estoque baixo/zerado + estimativa de custo de compra."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT
                s.nv_variant_id,
                s.nv_product_id,
                s.sku,
                s.quantity,
                s.min_quantity,
                sc.cost,
                sc.name  AS cost_name,
                sc.type  AS cost_type
            FROM sku_stock s
            LEFT JOIN sku_costs sc ON sc.id = (
                SELECT id FROM sku_costs sc2
                WHERE (
                    (s.nv_variant_id IS NOT NULL AND sc2.nv_variant_id = s.nv_variant_id)
                    OR (s.nv_variant_id IS NULL  AND sc2.sku = s.sku AND sc2.nv_variant_id IS NULL)
                )
                AND (sc2.effective_to IS NULL OR sc2.effective_to >= date('now'))
                ORDER BY sc2.effective_from DESC LIMIT 1
            )
            WHERE s.quantity <= s.min_quantity
            ORDER BY s.quantity ASC, s.sku ASC
        """).fetchall()

    items = []
    total_estimado = 0.0
    for r in rows:
        qty_atual  = r["quantity"]
        min_qty    = r["min_quantity"]
        custo      = float(r["cost"] or 0)
        # Quantidade sugerida para compra: repõe até 2× o mínimo
        qty_comprar = max(min_qty * 2 - qty_atual, min_qty, 1)
        total_item  = round(qty_comprar * custo, 2)
        total_estimado += total_item
        items.append({
            "nv_variant_id": r["nv_variant_id"],
            "sku":           r["sku"] or "—",
            "nome":          r["cost_name"] or r["sku"] or "—",
            "qty_atual":     qty_atual,
            "min_qty":       min_qty,
            "qty_comprar":   qty_comprar,
            "custo_unit":    custo,
            "total_item":    total_item,
            "status":        "sem_estoque" if qty_atual <= 0 else "estoque_baixo",
        })

    return jsonify({
        "items":          items,
        "total_estimado": round(total_estimado, 2),
        "sem_estoque":    sum(1 for i in items if i["status"] == "sem_estoque"),
        "estoque_baixo":  sum(1 for i in items if i["status"] == "estoque_baixo"),
        "total_itens":    len(items),
    })


@app.route("/api/alertas/poll")
@login_required
def poll_alertas():
    """Polling de alertas de estoque (substitui SSE para compatibilidade com Vercel)."""
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT sku, quantity, min_quantity, produto_nome, variante_label
            FROM sku_stock
            WHERE quantity <= min_quantity
            ORDER BY quantity ASC
            LIMIT 50
        """).fetchall()
    alertas = []
    for r in rows:
        alertas.append({
            "type":     "alerta_estoque",
            "status":   "sem_estoque" if (r["quantity"] or 0) <= 0 else "estoque_baixo",
            "produto":  r["produto_nome"] or r["sku"] or "—",
            "variante": r["variante_label"] or "",
            "sku":      r["sku"] or "",
            "qty_atual": r["quantity"] or 0,
            "min_qty":  r["min_quantity"] or 0,
            "timestamp": datetime.now().strftime("%H:%M:%S"),
        })
    return jsonify(alertas)


# ── Routes — Compras Manuais ──────────────────────────────────────────────────

@app.route("/api/compras/manual", methods=["GET"])
@login_required
def api_compras_manual_list():
    """Lista itens adicionados manualmente para o dia de hoje."""
    data_filtro = request.args.get("data", date.today().isoformat())
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM compras_manual WHERE data = ? ORDER BY created_at DESC",
            (data_filtro,)
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/compras/manual", methods=["POST"])
@login_required
def api_compras_manual_add():
    """Adiciona um item manualmente à lista de compras do dia."""
    d          = request.get_json() or {}
    nome       = str(d.get("nome", "")).strip()
    if not nome:
        return jsonify({"erro": "Nome é obrigatório"}), 400
    sku        = str(d.get("sku", "")).strip() or None
    qty        = max(1, int(d.get("qty_comprar", 1) or 1))
    custo      = max(0.0, float(d.get("custo_unit", 0) or 0))
    obs        = str(d.get("observacao", "")).strip() or None
    data_item  = d.get("data", date.today().isoformat())
    with get_conn() as conn:
        cur = conn.execute(
            """INSERT INTO compras_manual (nome, sku, qty_comprar, custo_unit, observacao, data)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (nome, sku, qty, custo, obs, data_item)
        )
        new_id = cur.lastrowid
    return jsonify({"ok": True, "id": new_id})


@app.route("/api/compras/manual/<int:item_id>", methods=["PUT"])
@login_required
def api_compras_manual_update(item_id):
    """Atualiza um item manual (nome, qty, custo, observação)."""
    d     = request.get_json() or {}
    nome  = str(d.get("nome", "")).strip()
    if not nome:
        return jsonify({"erro": "Nome é obrigatório"}), 400
    sku   = str(d.get("sku", "")).strip() or None
    qty   = max(1, int(d.get("qty_comprar", 1) or 1))
    custo = max(0.0, float(d.get("custo_unit", 0) or 0))
    obs   = str(d.get("observacao", "")).strip() or None
    with get_conn() as conn:
        conn.execute(
            """UPDATE compras_manual
               SET nome=?, sku=?, qty_comprar=?, custo_unit=?, observacao=?
               WHERE id=?""",
            (nome, sku, qty, custo, obs, item_id)
        )
    return jsonify({"ok": True})


@app.route("/api/compras/manual/<int:item_id>", methods=["DELETE"])
@login_required
def api_compras_manual_delete(item_id):
    """Remove um item manual."""
    with get_conn() as conn:
        conn.execute("DELETE FROM compras_manual WHERE id = ?", (item_id,))
    return jsonify({"ok": True})


# ── Routes — Compras Registros ────────────────────────────────────────────────

@app.route("/api/compras/registros", methods=["GET"])
@login_required
def api_compras_registros_list():
    """Lista compras registradas. Filtros: inicio, fim."""
    inicio = request.args.get("inicio") or ""
    fim    = request.args.get("fim")    or ""
    with get_conn() as conn:
        q = """
            SELECT r.*,
                   COALESCE(SUM(t.quantidade), 0) AS total_qty,
                   COALESCE(SUM(t.quantidade), 0) * r.preco_unit AS total_valor
            FROM compras_registros r
            LEFT JOIN compras_tamanhos t ON t.compra_id = r.id
        """
        params = []
        conds = []
        if inicio: conds.append("r.data >= ?"); params.append(inicio)
        if fim:    conds.append("r.data <= ?"); params.append(fim)
        if conds: q += " WHERE " + " AND ".join(conds)
        q += " GROUP BY r.id ORDER BY r.data DESC, r.created_at DESC"
        compras = [dict(r) for r in conn.execute(q, params).fetchall()]

        # Tamanhos de cada compra
        ids = [c["id"] for c in compras]
        tam_map = {}
        if ids:
            marks = ",".join("?" * len(ids))
            for t in conn.execute(
                f"SELECT * FROM compras_tamanhos WHERE compra_id IN ({marks}) ORDER BY tamanho",
                ids
            ).fetchall():
                tam_map.setdefault(t["compra_id"], []).append(dict(t))
        for c in compras:
            c["tamanhos"] = tam_map.get(c["id"], [])

    return jsonify(compras)


@app.route("/api/compras/registros", methods=["POST"])
@login_required
def api_compras_registros_add():
    """Registra uma nova compra com tamanhos."""
    data    = request.get_json() or {}
    produto = (data.get("produto_nome") or "").strip()
    if not produto:
        return jsonify({"erro": "Nome do produto obrigatório"}), 400

    tamanhos = [t for t in (data.get("tamanhos") or []) if int(t.get("quantidade") or 0) > 0]
    if not tamanhos:
        return jsonify({"erro": "Informe ao menos um tamanho com quantidade > 0"}), 400

    preco_unit       = float(data.get("preco_unit") or 0)
    atualizar_estoque= bool(data.get("atualizar_estoque"))
    agora            = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    hoje             = date.today().isoformat()
    data_compra      = data.get("data") or hoje
    estoque_atualizados = 0

    with get_conn() as conn:
        cur = conn.execute(
            """INSERT INTO compras_registros
               (data, produto_nome, nv_product_id, fornecedor, preco_unit, observacao, criado_por)
               VALUES (?,?,?,?,?,?,?)""",
            (data_compra, produto,
             data.get("nv_product_id") or None,
             (data.get("fornecedor") or "").strip() or None,
             preco_unit,
             (data.get("observacao") or "").strip() or None,
             session.get("usuario", ""))
        )
        cid = cur.lastrowid
        for t in tamanhos:
            conn.execute(
                "INSERT INTO compras_tamanhos (compra_id, tamanho, quantidade) VALUES (?,?,?)",
                (cid, str(t["tamanho"]).strip().upper(), int(t["quantidade"]))
            )

        # ── Atualiza estoque + CMP se solicitado ──────────────────────────
        if atualizar_estoque and preco_unit > 0:
            for t in tamanhos:
                nv_vid   = t.get("nv_variant_id") or None
                sku_tam  = t.get("sku") or None
                qty      = int(t["quantidade"])
                var_label= t.get("variante_label") or t["tamanho"]

                if not nv_vid and not sku_tam:
                    continue  # sem vínculo ao estoque

                # sku_custo = sku ou fallback VID
                sku_custo = sku_tam or (f"VID-{nv_vid}" if nv_vid else None)
                if not sku_custo:
                    continue

                # Estoque atual
                if nv_vid:
                    stock = conn.execute(
                        "SELECT id, quantity, produto_nome, variante_label FROM sku_stock WHERE nv_variant_id=?",
                        (nv_vid,)
                    ).fetchone()
                else:
                    stock = conn.execute(
                        "SELECT id, quantity, produto_nome, variante_label FROM sku_stock WHERE sku=? AND nv_variant_id IS NULL",
                        (sku_tam,)
                    ).fetchone()

                qty_atual = int(stock["quantity"]) if stock and stock["quantity"] else 0

                # Custo atual
                if nv_vid:
                    cr = conn.execute(
                        "SELECT cost FROM sku_costs WHERE nv_variant_id=? AND effective_to IS NULL ORDER BY effective_from DESC LIMIT 1",
                        (nv_vid,)
                    ).fetchone()
                else:
                    cr = conn.execute(
                        "SELECT cost FROM sku_costs WHERE sku=? AND nv_variant_id IS NULL AND effective_to IS NULL ORDER BY effective_from DESC LIMIT 1",
                        (sku_custo,)
                    ).fetchone()

                custo_atual = float(cr["cost"]) if cr else 0.0

                # CMP
                if qty_atual <= 0 or custo_atual == 0:
                    novo_cmp = preco_unit
                else:
                    novo_cmp = round((qty_atual * custo_atual + qty * preco_unit) / (qty_atual + qty), 4)

                nova_qty = qty_atual + qty

                # Atualiza sku_stock
                if stock:
                    conn.execute(
                        "UPDATE sku_stock SET quantity=?, updated_at=? WHERE id=?",
                        (nova_qty, agora, stock["id"])
                    )
                else:
                    conn.execute(
                        """INSERT INTO sku_stock (nv_variant_id, sku, quantity, produto_nome, variante_label, updated_at)
                           VALUES (?,?,?,?,?,?)""",
                        (nv_vid, sku_tam, nova_qty, produto, var_label, agora)
                    )

                # Movimento
                conn.execute(
                    """INSERT INTO sku_stock_movements
                       (nv_variant_id, sku, tipo, quantidade, preco_compra, observacao, created_at)
                       VALUES (?,?,?,?,?,?,?)""",
                    (nv_vid, sku_custo, "entrada", qty, preco_unit,
                     f"Compra registrada #{cid} — CMP: R${novo_cmp:.2f}", agora)
                )

                # Fecha custo anterior e insere CMP
                if nv_vid:
                    conn.execute(
                        "UPDATE sku_costs SET effective_to=date(?,' -1 day') WHERE nv_variant_id=? AND effective_to IS NULL AND effective_from < ?",
                        (hoje, nv_vid, hoje)
                    )
                conn.execute(
                    "UPDATE sku_costs SET effective_to=date(?,' -1 day') WHERE sku=? AND effective_to IS NULL AND effective_from < ?",
                    (hoje, sku_custo, hoje)
                )
                conn.execute(
                    """INSERT INTO sku_costs (sku, name, type, cost, effective_from, notes, nv_variant_id)
                       VALUES (?,?,?,?,?,?,?)""",
                    (sku_custo, produto, "product", novo_cmp, hoje,
                     f"CMP via compra #{cid}: {qty_atual}un×R${custo_atual:.2f} + {qty}un×R${preco_unit:.2f}",
                     nv_vid)
                )

                # Sync por nome+SKU
                _sync_estoque_por_nome(
                    conn, produto, var_label, nova_qty, agora,
                    skip_vid=nv_vid, sku_origem=sku_tam
                )

                estoque_atualizados += 1

        conn.commit()
    return jsonify({"ok": True, "id": cid, "estoque_atualizado": estoque_atualizados}), 201


@app.route("/api/compras/registros/<int:cid>", methods=["DELETE"])
@login_required
def api_compras_registros_delete(cid):
    with get_conn() as conn:
        conn.execute("DELETE FROM compras_tamanhos  WHERE compra_id=?", (cid,))
        conn.execute("DELETE FROM compras_registros WHERE id=?", (cid,))
        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/compras/relatorio", methods=["GET"])
@login_required
def api_compras_relatorio():
    """Relatório de compras: semanal e mensal."""
    with get_conn() as conn:
        # Por semana (últimas 8 semanas)
        semanal = conn.execute("""
            SELECT strftime('%Y-W%W', r.data) AS semana,
                   COUNT(DISTINCT r.id)       AS num_compras,
                   SUM(t.quantidade)          AS total_pecas,
                   SUM(t.quantidade * r.preco_unit) AS total_valor
            FROM compras_registros r
            LEFT JOIN compras_tamanhos t ON t.compra_id = r.id
            WHERE r.data >= date('now', '-56 days')
            GROUP BY semana
            ORDER BY semana DESC
        """).fetchall()

        # Por mês (últimos 12 meses)
        mensal = conn.execute("""
            SELECT strftime('%Y-%m', r.data)  AS mes,
                   COUNT(DISTINCT r.id)       AS num_compras,
                   SUM(t.quantidade)          AS total_pecas,
                   SUM(t.quantidade * r.preco_unit) AS total_valor
            FROM compras_registros r
            LEFT JOIN compras_tamanhos t ON t.compra_id = r.id
            WHERE r.data >= date('now', '-365 days')
            GROUP BY mes
            ORDER BY mes DESC
        """).fetchall()

        # Por produto (top 10)
        por_produto = conn.execute("""
            SELECT r.produto_nome,
                   COUNT(DISTINCT r.id)       AS num_compras,
                   SUM(t.quantidade)          AS total_pecas,
                   SUM(t.quantidade * r.preco_unit) AS total_valor
            FROM compras_registros r
            LEFT JOIN compras_tamanhos t ON t.compra_id = r.id
            GROUP BY UPPER(TRIM(r.produto_nome))
            ORDER BY total_valor DESC
            LIMIT 10
        """).fetchall()

        # Por fornecedor
        por_fornecedor = conn.execute("""
            SELECT COALESCE(r.fornecedor, 'Sem fornecedor') AS fornecedor,
                   COUNT(DISTINCT r.id) AS num_compras,
                   SUM(t.quantidade * r.preco_unit) AS total_valor
            FROM compras_registros r
            LEFT JOIN compras_tamanhos t ON t.compra_id = r.id
            GROUP BY UPPER(TRIM(COALESCE(r.fornecedor,'')))
            ORDER BY total_valor DESC
        """).fetchall()

    return jsonify({
        "semanal":       [dict(r) for r in semanal],
        "mensal":        [dict(r) for r in mensal],
        "por_produto":   [dict(r) for r in por_produto],
        "por_fornecedor":[dict(r) for r in por_fornecedor],
    })


# ── Routes — Atacado ──────────────────────────────────────────────────────────

@app.route("/atacado")
@login_required
def atacado_page():
    return render_template("atacado.html")


@app.route("/api/atacado/pedidos", methods=["GET"])
@login_required
def api_atacado_list():
    status_f = request.args.get("status", "")
    busca    = (request.args.get("q") or "").strip()
    pago_f   = request.args.get("pago", "")        # "1", "0" ou ""
    inicio   = (request.args.get("inicio") or "").strip()
    fim      = (request.args.get("fim") or "").strip()
    with get_conn() as conn:
        sql = """
            SELECT ap.*,
                   COUNT(ai.id)                        AS total_itens,
                   SUM(ai.quantidade)                  AS total_qty,
                   SUM(CASE WHEN ai.separado=1 THEN ai.quantidade ELSE 0 END) AS qty_separada,
                   SUM(COALESCE(ai.valor_unit,0) * ai.quantidade) AS valor_total
            FROM atacado_pedidos ap
            LEFT JOIN atacado_itens ai ON ai.pedido_id = ap.id
        """
        conds, params = [], []
        if status_f and status_f != "todos":
            conds.append("ap.status = ?"); params.append(status_f)
        if busca:
            conds.append("(ap.cliente LIKE ? OR ap.nome LIKE ? OR ap.contato LIKE ? OR CAST(ap.numero AS TEXT) LIKE ?)")
            like = f"%{busca}%"
            params += [like, like, like, like]
        if pago_f in ("0", "1"):
            conds.append("ap.pago = ?"); params.append(int(pago_f))
        if inicio:
            conds.append("date(ap.created_at) >= ?"); params.append(inicio)
        if fim:
            conds.append("date(ap.created_at) <= ?"); params.append(fim)
        if conds:
            sql += " WHERE " + " AND ".join(conds)
        sql += " GROUP BY ap.id ORDER BY ap.numero DESC, ap.created_at DESC"
        rows = conn.execute(sql, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/atacado/pedidos", methods=["POST"])
@login_required
def api_atacado_criar():
    data    = request.get_json() or {}
    cliente = (data.get("cliente") or "").strip()
    contato = (data.get("contato") or "").strip()
    if not cliente:
        return jsonify({"erro": "Nome do cliente é obrigatório"}), 400
    if not contato:
        return jsonify({"erro": "Número de WhatsApp é obrigatório"}), 400
    itens = data.get("itens") or []
    if not itens:
        return jsonify({"erro": "Adicione ao menos um item"}), 400

    with get_conn() as conn:
        # Número sequencial — começa em 1
        proximo_numero = (conn.execute(
            "SELECT COALESCE(MAX(numero), 0) FROM atacado_pedidos"
        ).fetchone()[0] or 0) + 1

        cur = conn.execute(
            """INSERT INTO atacado_pedidos
               (numero, cliente, nome, contato, cep, endereco, cidade, cpf,
                observacao, prazo, pago, frete_tipo, criado_por)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (proximo_numero, cliente,
             data.get("nome",""), contato,
             data.get("cep",""),  data.get("endereco",""),
             data.get("cidade",""), data.get("cpf",""),
             data.get("observacao",""), data.get("prazo",""),
             1 if data.get("pago") else 0,
             data.get("frete_tipo","a_combinar") or "a_combinar",
             session.get("usuario",""))
        )
        pedido_id = cur.lastrowid
        for item in itens:
            produto   = (item.get("produto") or "").strip()
            variante  = (item.get("variante") or "").strip()
            qty_est   = max(int(item.get("qty_estoque") or 0), 0)
            qty_forn  = max(int(item.get("qty_fornecedor") or 0), 0)
            qty_total = qty_est + qty_forn
            nv_vid    = item.get("nv_variant_id") or None
            valor_unit = float(item.get("valor_unit") or 0)
            if produto and qty_total > 0:
                conn.execute(
                    """INSERT INTO atacado_itens
                       (pedido_id, produto, variante, quantidade,
                        qty_estoque, qty_fornecedor, nv_variant_id, valor_unit)
                       VALUES (?,?,?,?,?,?,?,?)""",
                    (pedido_id, produto, variante, qty_total,
                     qty_est, qty_forn, nv_vid, valor_unit)
                )
        conn.commit()
    return jsonify({"ok": True, "id": pedido_id, "numero": proximo_numero}), 201


@app.route("/api/atacado/pedidos/<int:pid>", methods=["GET"])
@login_required
def api_atacado_detalhe(pid):
    with get_conn() as conn:
        ped  = conn.execute("SELECT * FROM atacado_pedidos WHERE id=?", (pid,)).fetchone()
        if not ped:
            return jsonify({"erro": "Pedido não encontrado"}), 404
        itens = conn.execute(
            "SELECT * FROM atacado_itens WHERE pedido_id=? ORDER BY id", (pid,)
        ).fetchall()
    return jsonify({"pedido": dict(ped), "itens": [dict(i) for i in itens]})


@app.route("/api/atacado/pedidos/<int:pid>", methods=["DELETE"])
@login_required
def api_atacado_deletar(pid):
    with get_conn() as conn:
        conn.execute("DELETE FROM atacado_itens WHERE pedido_id=?", (pid,))
        conn.execute("DELETE FROM atacado_pedidos WHERE id=?", (pid,))
        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/atacado/itens/<int:iid>/separado", methods=["PUT"])
@login_required
def api_atacado_item_separado(iid):
    data     = request.get_json() or {}
    separado = 1 if data.get("separado") else 0
    agora    = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    usuario  = session.get("usuario", "")
    with get_conn() as conn:
        conn.execute(
            """UPDATE atacado_itens
               SET separado=?, separado_em=?, separado_por=?
               WHERE id=?""",
            (separado, agora if separado else None,
             usuario if separado else None, iid)
        )
        # Atualiza status do pedido automaticamente
        ped_row = conn.execute(
            "SELECT pedido_id FROM atacado_itens WHERE id=?", (iid,)
        ).fetchone()
        if ped_row:
            pid = ped_row["pedido_id"]
            total  = conn.execute("SELECT COUNT(*) FROM atacado_itens WHERE pedido_id=?", (pid,)).fetchone()[0]
            sep    = conn.execute("SELECT COUNT(*) FROM atacado_itens WHERE pedido_id=? AND separado=1", (pid,)).fetchone()[0]
            status_atual = conn.execute("SELECT status FROM atacado_pedidos WHERE id=?", (pid,)).fetchone()["status"]
            if status_atual not in ("enviado",):
                novo_status = "separado" if sep == total and total > 0 else ("separando" if sep > 0 else "pendente")
                conn.execute("UPDATE atacado_pedidos SET status=?, updated_at=? WHERE id=?",
                             (novo_status, agora, pid))
        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/atacado/pedidos/<int:pid>/enviado", methods=["PUT"])
@login_required
def api_atacado_marcar_enviado(pid):
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        conn.execute("UPDATE atacado_pedidos SET status='enviado', updated_at=? WHERE id=?", (agora, pid))
        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/atacado/pedidos/<int:pid>/pago", methods=["PUT"])
@login_required
def api_atacado_pago(pid):
    data = request.get_json() or {}
    pago = 1 if data.get("pago") else 0
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        conn.execute("UPDATE atacado_pedidos SET pago=?, updated_at=? WHERE id=?", (pago, agora, pid))
        conn.commit()
    return jsonify({"ok": True})


def _descontar_item_estoque(conn, item, agora):
    """Tenta descontar um item do estoque. Retorna (sucesso, motivo).
    NÃO desconta se faltar estoque — só desconta o que existe de verdade."""
    if item["estoque_descontado"]:
        return False, "ja_descontado"

    qty_est = item["qty_estoque"] or 0
    nv_vid  = item["nv_variant_id"]

    # Item sem quantidade do estoque (tudo do fornecedor) → marca como ok, nada a descontar
    if qty_est <= 0:
        conn.execute("UPDATE atacado_itens SET estoque_descontado=1 WHERE id=?", (item["id"],))
        return True, "sem_qty_estoque"

    # Item sem vínculo de estoque (não está no catálogo de estoque)
    if not nv_vid:
        return False, "sem_vinculo"

    stock = conn.execute(
        "SELECT id, quantity, produto_nome, variante_label, sku FROM sku_stock WHERE nv_variant_id=?",
        (nv_vid,)
    ).fetchone()

    if not stock:
        return False, "sem_cadastro"

    disponivel = stock["quantity"] or 0
    if disponivel < qty_est:
        # Estoque insuficiente — NÃO desconta, avisa
        return False, f"insuficiente:{disponivel}:{qty_est}"

    # OK — desconta
    nova_qty = disponivel - qty_est
    conn.execute("UPDATE sku_stock SET quantity=?, updated_at=? WHERE id=?",
                 (nova_qty, agora, stock["id"]))
    conn.execute("""INSERT INTO sku_stock_movements
        (nv_variant_id, sku, tipo, quantidade, pedido_numero, observacao, created_at)
        VALUES (?,?,?,?,?,?,?)""",
        (nv_vid, item["variante"] or "", "saida_venda",
         qty_est, f"ATK-{item['pedido_id']}", "Saída por pedido de atacado", agora))
    _deducao_sync_por_nome(
        conn, stock["produto_nome"], stock["variante_label"],
        qty_est, agora, skip_vid=nv_vid,
        pedido_nr=f"ATK-{item['pedido_id']}",
        sku_origem=stock["sku"]
    )
    conn.execute("UPDATE atacado_itens SET estoque_descontado=1 WHERE id=?", (item["id"],))
    return True, "descontado"


@app.route("/api/atacado/itens/<int:iid>/descontar-estoque", methods=["PUT"])
@login_required
def api_atacado_descontar_estoque(iid):
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        item = conn.execute("SELECT * FROM atacado_itens WHERE id=?", (iid,)).fetchone()
        if not item:
            return jsonify({"erro": "Item não encontrado"}), 404
        if item["estoque_descontado"]:
            return jsonify({"erro": "Estoque já foi descontado"}), 400

        ok, motivo = _descontar_item_estoque(conn, item, agora)
        if not ok:
            if motivo.startswith("insuficiente:"):
                _, disp, prec = motivo.split(":")
                return jsonify({"erro": f"Estoque insuficiente: há {disp} em estoque, mas o pedido precisa de {prec}."}), 400
            if motivo == "sem_cadastro":
                return jsonify({"erro": "Este produto não tem estoque cadastrado."}), 400
            if motivo == "sem_vinculo":
                return jsonify({"erro": "Item sem vínculo com o catálogo de estoque."}), 400
            return jsonify({"erro": "Não foi possível descontar."}), 400

        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/atacado/pedidos/<int:pid>/descontar-tudo", methods=["PUT"])
@login_required
def api_atacado_descontar_tudo(pid):
    """Desconta TODOS os itens do pedido do estoque de uma vez.
    Itens sem estoque suficiente são pulados e reportados — não bloqueiam os outros."""
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        ped = conn.execute("SELECT id FROM atacado_pedidos WHERE id=?", (pid,)).fetchone()
        if not ped:
            return jsonify({"erro": "Pedido não encontrado"}), 404

        itens = conn.execute(
            "SELECT * FROM atacado_itens WHERE pedido_id=? ORDER BY id", (pid,)
        ).fetchall()

        descontados = 0
        problemas   = []   # itens que não puderam ser descontados

        for item in itens:
            if item["estoque_descontado"]:
                continue
            if (item["qty_estoque"] or 0) <= 0:
                # nada do estoque — marca como tratado, sem descontar
                conn.execute("UPDATE atacado_itens SET estoque_descontado=1 WHERE id=?", (item["id"],))
                continue

            ok, motivo = _descontar_item_estoque(conn, item, agora)
            nome_item = item["produto"] + (f" — {item['variante']}" if item["variante"] else "")
            if ok:
                descontados += 1
            else:
                if motivo.startswith("insuficiente:"):
                    _, disp, prec = motivo.split(":")
                    problemas.append({"produto": nome_item, "motivo": f"Só há {disp} em estoque (precisa {prec})"})
                elif motivo == "sem_cadastro":
                    problemas.append({"produto": nome_item, "motivo": "Sem estoque cadastrado"})
                elif motivo == "sem_vinculo":
                    problemas.append({"produto": nome_item, "motivo": "Sem vínculo com o estoque"})

        conn.commit()

    return jsonify({
        "ok": True,
        "descontados": descontados,
        "problemas": problemas,
    })


@app.route("/api/atacado/pedidos/<int:pid>/reabrir", methods=["PUT"])
@login_required
def api_atacado_reabrir(pid):
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        total = conn.execute("SELECT COUNT(*) FROM atacado_itens WHERE pedido_id=?", (pid,)).fetchone()[0]
        sep   = conn.execute("SELECT COUNT(*) FROM atacado_itens WHERE pedido_id=? AND separado=1", (pid,)).fetchone()[0]
        novo  = "separado" if sep == total and total > 0 else ("separando" if sep > 0 else "pendente")
        conn.execute("UPDATE atacado_pedidos SET status=?, updated_at=? WHERE id=?", (novo, agora, pid))
        conn.commit()
    return jsonify({"ok": True})


# ── /links removido (dependia de arquivos locais do Windows) ─────────────────


# ── Entry point ───────────────────────────────────────────────────────────────

# Inicializa o banco na primeira importação (cold start do Vercel)
try:
    init_db()
except Exception:
    pass


# ── NuvemShop LGPD webhooks ───────────────────────────────────────────────────

@app.route("/nuvemshop/webhooks/store-redact", methods=["POST"])
def lgpd_store_redact():
    return jsonify({"ok": True}), 200

@app.route("/nuvemshop/webhooks/customers-redact", methods=["POST"])
def lgpd_customers_redact():
    return jsonify({"ok": True}), 200

@app.route("/nuvemshop/webhooks/customers-data-request", methods=["POST"])
def lgpd_customers_data_request():
    return jsonify({"ok": True}), 200


if __name__ == "__main__":
    app.run(debug=False, host="0.0.0.0", port=5000)
